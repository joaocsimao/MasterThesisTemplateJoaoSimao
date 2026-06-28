#!/usr/bin/env python3
"""
run_qwk_simple.py
-----------------
Finds all MASGraded.csv files under a root directory and produces
qwk_results.xlsx with 2 sheets:

  Overall       – AI×score metrics (all rows, per CSV)
  Interpretation – metric reference guide

Usage:
    python run_qwk_simple.py [root_dir]            (default: current directory)
    python run_qwk_simple.py [root_dir] --skip-rows

Flags:
    --skip-rows   Collect all row identifiers (column: 'identifier') that are
                  skipped (AI_grade == -1) across ALL CSVs, then restrict every
                  metric sheet to only the rows whose identifier never appears
                  in that skipped set.

Expected CSV columns:
    score, AI_grade, identifier  (identifier required when --skip-rows is used)
"""

import sys, os, glob, argparse
import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from scipy.stats import pearsonr, spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Metric helpers ─────────────────────────────────────────────────────────────

def _mask(y1, y2):
    m = y1.notna() & y2.notna()
    return y1[m], y2[m], int(m.sum())

def qwk(y1, y2):
    a, b, n = _mask(y1, y2)
    if n < 2: return float("nan")
    try:
        return cohen_kappa_score(a.astype(int), b.astype(int), weights="quadratic")
    except Exception:
        return float("nan")

def exact_acc(y1, y2):
    a, b, n = _mask(y1, y2)
    return float("nan") if n == 0 else (a.astype(int) == b.astype(int)).mean()

def adj_acc(y1, y2):
    a, b, n = _mask(y1, y2)
    return float("nan") if n == 0 else (abs(a.astype(int) - b.astype(int)) <= 1).mean()

def mae(y1, y2):
    a, b, n = _mask(y1.astype(float), y2.astype(float))
    return float("nan") if n == 0 else (a - b).abs().mean()

def rmse(y1, y2):
    a, b, n = _mask(y1.astype(float), y2.astype(float))
    return float("nan") if n == 0 else np.sqrt(((a - b) ** 2).mean())

def pearson(y1, y2):
    a, b, n = _mask(y1.astype(float), y2.astype(float))
    if n < 2: return float("nan")
    r, _ = pearsonr(a, b); return r

def spearman(y1, y2):
    a, b, n = _mask(y1.astype(float), y2.astype(float))
    if n < 2: return float("nan")
    rho, _ = spearmanr(a, b); return rho

def all_metrics(y1, y2):
    _, _, n = _mask(y1, y2)
    def s(v): return None if (v is None or (isinstance(v, float) and np.isnan(v))) else round(float(v), 4)
    return {
        "n":        n,
        "qwk":      s(qwk(y1, y2)),
        "exact":    s(exact_acc(y1, y2)),
        "adj":      s(adj_acc(y1, y2)),
        "mae":      s(mae(y1, y2)),
        "rmse":     s(rmse(y1, y2)),
        "pearson":  s(pearson(y1, y2)),
        "spearman": s(spearman(y1, y2)),
    }

# ── CSV loading ────────────────────────────────────────────────────────────────

def _read_raw(path):
    """Read a CSV and normalise column names. No filtering applied."""
    df = pd.read_csv(
        path,
        sep=",",
        encoding="utf-8",
        encoding_errors="replace",
        low_memory=False,
    )
    df.columns = df.columns.str.strip()
    return df


def load_csv(path):
    """
    Load and clean a CSV for metric computation.

    Returns (df_clean, skipped_count, skipped_identifiers):
        df_clean            – rows ready for metrics (sentinel -1 rows removed)
        skipped_count       – number of AI_grade == -1 rows removed
        skipped_identifiers – set of identifier values from those removed rows
                              (empty set if 'identifier' column is absent)
    """
    df = _read_raw(path)

    # Remove token usage / summary rows
    if "score" in df.columns:
        df = df[
            ~df["score"].astype(str).str.contains(
                "token|summary|usage|total",
                case=False,
                na=False,
            )
        ].reset_index(drop=True)

    skipped_count = 0
    skipped_identifiers: set = set()

    if "AI_grade" in df.columns:
        df["AI_grade"] = pd.to_numeric(
            df["AI_grade"].astype(str).str.strip(),
            errors="coerce",
        )

        sentinel_mask = df["AI_grade"] == -1
        skipped_count = int(sentinel_mask.sum())

        # Capture identifiers of skipped rows before dropping them
        if skipped_count and "identifier" in df.columns:
            skipped_identifiers = set(df.loc[sentinel_mask, "identifier"].dropna().tolist())

        df = df[~sentinel_mask].reset_index(drop=True)

    for col in ["score", "AI_grade", "essay_word_count"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df, skipped_count, skipped_identifiers


def build_global_skip_set(files):
    """
    Collect, across ALL files, the identifier of every row where
    AI_grade == -1.  Any identifier that appears in this set in even
    one file will be excluded from every file before metrics are run.

    Returns a set of identifier values.
    """
    global_skipped: set = set()
    missing_col = []

    for path in files:
        _, _, ids = load_csv(path)
        if ids:
            global_skipped.update(ids)
        else:
            # Check whether the column is simply absent (vs. no -1 rows)
            df_raw = _read_raw(path)
            if "identifier" not in df_raw.columns:
                missing_col.append(path)

    if missing_col:
        print(f"  ⚠  --skip-rows: 'identifier' column missing in {len(missing_col)} file(s):")
        for p in missing_col:
            print(f"       {p}")

    print(f"  --skip-rows: {len(global_skipped)} unique identifier(s) flagged across all files.\n")
    return global_skipped


def valid_mask(df, skipped_ids):
    """
    Return a boolean Series that is True for rows whose identifier is NOT
    in the global skip set.  If --skip-rows was not used (skipped_ids empty)
    or the column is absent, every row is considered valid.
    The DataFrame is never modified.
    """
    if not skipped_ids or "identifier" not in df.columns:
        return pd.Series(True, index=df.index)
    return ~df["identifier"].isin(skipped_ids)


def parent_name(path):
    parts = path.replace("\\", "/").split("/")
    parents = parts[-3:-1] if len(parts) >= 3 else parts[:-1]
    return " / ".join(parents)

def filter_by_word_count(df, mode):
    """
    mode:
        under_100  -> essay_word_count < 100
        between    -> 100 <= essay_word_count <= 400
        over_400   -> essay_word_count > 400
    """
    if "essay_word_count" not in df.columns:
        return pd.DataFrame()

    wc = pd.to_numeric(df["essay_word_count"], errors="coerce")

    if mode == "under_100":
        return df[wc < 200]
    elif mode == "between":
        return df[(wc >= 100) & (wc <= 400)]
    elif mode == "over_400":
        return df[wc > 400]

    return pd.DataFrame()

# ── Excel style helpers ────────────────────────────────────────────────────────

THIN   = Side(style="thin",   color="B8CCE4")
MED    = Side(style="medium", color="1F4E79")
BORD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORD_T = Border(left=THIN, right=THIN, top=MED,  bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")

def _fill(color): return PatternFill("solid", start_color=color)

def hdr(ws, row, col, value, bg="1F4E79", fg="FFFFFF", size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=fg, size=size)
    c.fill = _fill(bg); c.alignment = CENTER; c.border = BORD
    return c

def dat(ws, row, col, value, even=True, bold=False, top=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.fill = _fill("DEEAF1" if even else "FFFFFF")
    c.border = BORD_T if top else BORD
    if isinstance(value, (int, float)) and value is not None:
        try:
            if not np.isnan(float(value)):
                c.alignment = CENTER
                c.number_format = "0.0000" if isinstance(value, float) else "0"
            else:
                c.alignment = CENTER
        except Exception:
            c.alignment = CENTER
    else:
        c.alignment = LEFT
    return c

# ── Sheet writer ───────────────────────────────────────────────────────────────

METRIC_COLS = [
    ("N",          7),  ("Skipped",    10), ("QWK",        9),
    ("Exact Acc", 11),  ("±1 Acc",      9), ("MAE",         9),
    ("RMSE",       9),  ("Pearson r",  10), ("Spearman ρ", 11),
]

def write_metrics_sheet(wb, rows, sheet_name, title):
    ws = wb.create_sheet(sheet_name)

    fixed = [("Name (Parent1 / Parent2)", 32), ("Rater A", 14), ("Rater B", 14)]
    all_cols = fixed + METRIC_COLS
    ncols = len(all_cols)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = _fill("10375C"); c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Group sub-headers
    groups = [("Identification", 1, 3), ("Count", 4, 5), ("Agreement Metrics", 6, ncols)]
    for label, c1, c2 in groups:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        hdr(ws, 2, c1, label, bg="2E75B6")
        for cx in range(c1 + 1, c2 + 1):
            ws.cell(row=2, column=cx).fill = _fill("2E75B6")
            ws.cell(row=2, column=cx).border = BORD
    ws.row_dimensions[2].height = 16

    # Column headers + widths
    for i, (h, w) in enumerate(all_cols, 1):
        hdr(ws, 3, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    prev_name = None
    color_flip = True
    excel_row = 4

    for r in rows:
        name_key = r.get("name", "")
        if name_key != prev_name:
            color_flip = not color_flip
            top_border = prev_name is not None
            prev_name = name_key
        else:
            top_border = False

        vals = [
            r.get("name", ""), r["rater_a"], r["rater_b"],
            r["n"], r.get("skipped", ""),
            r["qwk"], r["exact"], r["adj"],
            r["mae"], r["rmse"], r["pearson"], r["spearman"],
        ]
        for col, v in enumerate(vals, 1):
            dat(ws, excel_row, col, v, even=color_flip, top=top_border and col == 1)
            if top_border:
                ws.cell(row=excel_row, column=col).border = BORD_T
        excel_row += 1

# ── Interpretation sheet ───────────────────────────────────────────────────────

def write_interp_sheet(wb):
    ws = wb.create_sheet("Interpretation")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    rows = [
        ("Metric / Term",  "Explanation"),
        ("QWK",            "Quadratic Weighted Kappa. Measures ordinal agreement penalising larger disagreements more. "
                           "<0: Poor | 0–0.20: Slight | 0.21–0.40: Fair | 0.41–0.60: Moderate | 0.61–0.80: Substantial | 0.81–1.00: Almost perfect"),
        ("Exact Accuracy", "Proportion of grades matching exactly. Range 0–1; higher is better."),
        ("±1 Accuracy",    "Proportion of grades within 1 point of each other. Practical tolerance — captures near-misses acceptable in human grading."),
        ("MAE",            "Mean Absolute Error — average absolute grade difference. Lower is better."),
        ("RMSE",           "Root Mean Squared Error — like MAE but penalises large errors more heavily. Lower is better."),
        ("Pearson r",      "Linear correlation between grade series. >0.70 strong, >0.90 very strong."),
        ("Spearman ρ",     "Rank correlation — same thresholds as Pearson. More robust when distributions are skewed."),
        ("Skipped",        "Rows where AI_grade == -1 (sentinel / ungraded). Excluded before computing all metrics."),
        ("Overall sheet",  "AI agreement metrics computed across ALL rows combined (score × AI_grade)."),
        ("Name column",    "Derived from the two parent directories of the MASGraded.csv file path (e.g. ModelGPT4 / PortugalRound1)."),
        ("--skip-rows",    "When this flag is used, any identifier that has AI_grade == -1 in ANY CSV is collected into a "
                           "global skip set. That identifier is then removed from every CSV before metrics are computed, "
                           "ensuring all sheets compare only rows that were successfully graded everywhere."),
    ]
    for r_idx, (a, b) in enumerate(rows, 1):
        ca = ws.cell(row=r_idx, column=1, value=a)
        cb = ws.cell(row=r_idx, column=2, value=b)
        if r_idx == 1:
            for c in (ca, cb):
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill = _fill("1F4E79"); c.alignment = CENTER; c.border = BORD
        else:
            bg = "DEEAF1" if r_idx % 2 == 0 else "FFFFFF"
            for c in (ca, cb):
                c.font = Font(name="Arial", size=10); c.fill = _fill(bg); c.border = BORD
            ca.alignment = LEFT
            cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 36

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Compute AI grading agreement metrics from MASGraded.csv files."
    )
    parser.add_argument(
        "root",
        nargs="?",
        default=".",
        help="Root directory to search for MASGraded.csv files (default: current directory)"
    )
    parser.add_argument(
        "--skip-rows",
        action="store_true",
        help=(
            "Collect identifiers (column: 'identifier') of all rows with "
            "AI_grade == -1 across ALL CSVs, then exclude those identifiers "
            "from every file before computing metrics."
        )
    )
    args = parser.parse_args()
    root = args.root

    files = sorted(glob.glob(os.path.join(root, "**", "MASGraded.csv"), recursive=True))
    if not files:
        sys.exit(f"No MASGraded.csv files found under '{root}'")
    print(f"Found {len(files)} MASGraded.csv file(s)\n")

    # ── First pass: build global skip set (only when flag is set) ─────────────
    skipped_ids: set = set()
    if args.skip_rows:
        print("--skip-rows enabled: scanning all files for skipped identifiers…")
        skipped_ids = build_global_skip_set(files)

    # ── Second pass: compute metrics ──────────────────────────────────────────
    overall_rows   = []
    under_100_rows = []
    between_rows   = []
    over_400_rows  = []

    for path in files:
        name = parent_name(path)
        print(f"Processing: {path}  →  {name}")
        df, skipped, _ = load_csv(path)
        print(f"  rows after sentinel filter: {len(df)}  |  skipped (AI_grade==-1): {skipped}")

        if "score" not in df.columns or "AI_grade" not in df.columns:
            print(f"  ⚠ Skipping — missing 'score' or 'AI_grade' column")
            continue

        # Build a boolean mask for rows not in the global skip set.
        # The DataFrame itself is never modified.
        mask = valid_mask(df, skipped_ids)
        excluded = int((~mask).sum())
        if excluded:
            print(f"  --skip-rows: excluding {excluded} row(s) whose identifier was flagged globally")

        # ── Overall ─────────────────────────────
        m = all_metrics(df.loc[mask, "score"], df.loc[mask, "AI_grade"])
        overall_rows.append({
            "name":    name,
            "file":    path,
            "rater_a": "score",
            "rater_b": "AI_grade",
            "skipped": skipped,
            **m,
        })

        # ── Under 100 ───────────────────────────
        df_under = filter_by_word_count(df[mask], "under_100")
        if not df_under.empty:
            m_under = all_metrics(df_under["score"], df_under["AI_grade"])
            under_100_rows.append({
                "name":    name,
                "file":    path,
                "rater_a": "score",
                "rater_b": "AI_grade",
                "skipped": skipped,
                **m_under,
            })

        # ── 100–400 ─────────────────────────────
        df_between = filter_by_word_count(df[mask], "between")
        if not df_between.empty:
            m_between = all_metrics(df_between["score"], df_between["AI_grade"])
            between_rows.append({
                "name":    name,
                "file":    path,
                "rater_a": "score",
                "rater_b": "AI_grade",
                "skipped": skipped,
                **m_between,
            })

        # ── Over 400 ────────────────────────────
        df_over = filter_by_word_count(df[mask], "over_400")
        if not df_over.empty:
            m_over = all_metrics(df_over["score"], df_over["AI_grade"])
            over_400_rows.append({
                "name":    name,
                "file":    path,
                "rater_a": "score",
                "rater_b": "AI_grade",
                "skipped": skipped,
                **m_over,
            })

        print(f"  ✓ Overall QWK={m['qwk']}  n={m['n']}")

    if not overall_rows:
        sys.exit("No metrics produced — check that CSVs contain 'score' and 'AI_grade' columns.")

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_metrics_sheet(
        wb,
        overall_rows,
        "Overall",
        "AI Agreement — score × AI_grade (all rows)"
    )
    write_metrics_sheet(
        wb,
        under_100_rows,
        "QWK_Under_100",
        "AI Agreement — essay_word_count < 100"
    )
    write_metrics_sheet(
        wb,
        between_rows,
        "QWK_100_400",
        "AI Agreement — 100 ≤ essay_word_count ≤ 400"
    )
    write_metrics_sheet(
        wb,
        over_400_rows,
        "QWK_Over_400",
        "AI Agreement — essay_word_count > 400"
    )
    write_interp_sheet(wb)

    out = os.path.join(".", "qwk_results.xlsx")
    wb.save(out)
    print(f"\nExcel saved → {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()