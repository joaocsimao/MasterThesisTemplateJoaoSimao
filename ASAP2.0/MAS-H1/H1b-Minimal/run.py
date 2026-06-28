from __future__ import annotations

import argparse
import csv
import json
import os
import socket
import threading
import time
import urllib.error
import urllib.request
from collections import deque
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent.parent
THIS_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / "asap2_total_master.csv"
OUTPUT_CSV = THIS_DIR / "MASGraded.csv"
SUMMARY_XLSX = BASE_DIR / "GradingSummary.xlsx"

RUBRIC_FILE = BASE_DIR / "Criteria.txt"

DEFAULT_API_VERSION = "2024-12-01-preview"
DEFAULT_TEMPERATURE = 0.0
DEFAULT_MAX_TOKENS_EXTRACT = 1000
DEFAULT_MAX_TOKENS_SCORE = 1000
DEFAULT_WORKERS = 150
ADAPTIVE_GROWTH_STREAK = 5
ROW_RETRY_LIMIT = 3
BASE_RETRY_DELAY_SECONDS = 2.0
MAX_RETRY_DELAY_SECONDS = 20.0
SUMMARY_HEADERS = [
    "Test Name",
    "Start Time",
    "End Time",
    "Duration",
    "Sequential Time",
    "Total Wall Seconds",
    "Speedup Factor",
    "Agent 1 Prompt Tokens",
    "Agent 1 Completion Tokens",
    "Agent 2 Prompt Tokens",
    "Agent 2 Completion Tokens",
    "Input Tokens",
    "Output Tokens",
    "Total Tokens",
    "Agent 1 Errors",
    "Agent 2 Errors",
    "Total Errors",
    "Skipped Rows",
    "Skipped Row IDs",
    "Start Workers",
    "Peak Workers",
    "Min Workers",
    "Retryable Events",
    "Error Rows",
]


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def get_env(name: str, default: str | None = None) -> str:
    value = os.getenv(name, default)
    if value is None or not value.strip():
        raise SystemExit(f"Missing required environment variable: {name}")
    return value.strip()


def get_rubric() -> str:
    if not RUBRIC_FILE.exists():
        raise SystemExit(f"Rubric file not found: {RUBRIC_FILE}")
    return load_text(RUBRIC_FILE)


def parse_retry_after(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


class AzureRetryableError(RuntimeError):
    def __init__(self, message: str, *, kind: str, retry_after: float | None = None) -> None:
        super().__init__(message)
        self.kind = kind
        self.retry_after = retry_after


class AzureThrottleError(AzureRetryableError):
    def __init__(self, message: str, *, retry_after: float | None = None) -> None:
        super().__init__(message, kind="throttle", retry_after=retry_after)


class AzureTransientError(AzureRetryableError):
    def __init__(self, message: str, *, retry_after: float | None = None) -> None:
        super().__init__(message, kind="transient", retry_after=retry_after)


def resolve_worker_count() -> int:
    raw = os.getenv("MAX_WORKERS", "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return DEFAULT_WORKERS


def resolve_growth_threshold() -> int:
    raw = os.getenv("ADAPTIVE_GROWTH_STREAK", "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return ADAPTIVE_GROWTH_STREAK


def retry_delay_seconds(attempt: int, retry_after: float | None = None) -> float:
    if retry_after is not None and retry_after > 0:
        return min(retry_after, MAX_RETRY_DELAY_SECONDS)
    delay = BASE_RETRY_DELAY_SECONDS * (2 ** attempt)
    return min(delay, MAX_RETRY_DELAY_SECONDS)


def call_azure_openai(
    messages: list[dict[str, str]],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
    max_tokens: int,
) -> tuple[str | None, int, int]:
    """Returns (content, prompt_tokens, completion_tokens). Returns (None, 0, 0) on content-filter hit."""
    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    payload = {
        "messages": messages,
        "temperature": DEFAULT_TEMPERATURE,
        "max_tokens": max_tokens,
        "response_format": {"type": "json_object"},
    }
    request = urllib.request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "api-key": api_key},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            response_body = response.read().decode("utf-8")
    except TimeoutError as exc:
        raise AzureTransientError(
            f"Azure OpenAI request timed out after 120 seconds: {exc}"
        ) from exc
    except socket.timeout as exc:
        raise AzureTransientError(
            f"Azure OpenAI request timed out after 120 seconds: {exc}"
        ) from exc
    except urllib.error.URLError as exc:
        raise AzureTransientError(f"Azure OpenAI request failed: {exc}") from exc
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        if "content_filter" in error_body:
            print("  Content filter triggered. Skipping this row.")
            return None, 0, 0
        if exc.code == 429:
            raise AzureThrottleError(
                f"Azure OpenAI request throttled: {exc.code} {exc.reason}\n{error_body}",
                retry_after=parse_retry_after(exc.headers.get("Retry-After")),
            ) from exc
        if exc.code in {408, 500, 502, 503, 504}:
            raise AzureTransientError(
                f"Azure OpenAI request failed transiently: {exc.code} {exc.reason}\n{error_body}"
            ) from exc
        raise RuntimeError(
            f"Azure OpenAI request failed: {exc.code} {exc.reason}\n{error_body}"
        ) from exc

    response_json = json.loads(response_body)
    usage = response_json.get("usage", {})
    prompt_tokens = usage.get("prompt_tokens", 0)
    completion_tokens = usage.get("completion_tokens", 0)

    choices = response_json.get("choices", [])
    if not choices:
        raise RuntimeError(f"Azure OpenAI returned no choices: {response_body}")

    content = choices[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"Azure OpenAI returned empty content: {response_body}")

    return content, prompt_tokens, completion_tokens

FEXTRACT_SYSTEM = (
    "You will be producing an output, that will be read by this agent profile:"
    "{You evaluate student work against a holistic standard. "
    "You receive a pre-synthesised briefing that flags exactly where scrutiny is needed. "
    "Follow the output instructions exactly.}"
    
)
FEXTRACT_OUTPUT_INSTRUCTIONS = (
      "create a json with the asnwer and the rubric."
)



FEXTRACT_USER_TEMPLATE = """\
Question: {question}

Rubric:
{rubric}

Student Response:
{student_answer}

Output Instructions: {output_instructions}
"""


def build_fextract_messages(
    question: str, student_answer: str, rubric: str, retry_note: str = ""
) -> list[dict[str, str]]:
    user_content = FEXTRACT_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        output_instructions=FEXTRACT_OUTPUT_INSTRUCTIONS,
    )
    if retry_note:
        user_content += f"\n\nIMPORTANT: {retry_note}"
    return [
        {"role": "system", "content": FEXTRACT_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def run_fextract(
    question: str,
    student_answer: str,
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any] | None, int, int, int, str | None]:
    retry_note = ""
    last_error: Exception | None = None
    total_pt = 0
    total_ct = 0
    error_count = 0

    for attempt in range(2):
        messages = build_fextract_messages(question, student_answer, rubric, retry_note)
        try:
            raw, pt, ct = call_azure_openai(
                messages,
                endpoint,
                api_key,
                deployment,
                api_version,
                max_tokens=DEFAULT_MAX_TOKENS_EXTRACT,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                break
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(
                f"  [fextract {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s..."
            )
            time.sleep(wait_seconds)
            retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only a valid JSON object with no markdown or extra text."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            return None, total_pt, total_ct, error_count, None

        try:
            parsed = json.loads(raw)
            if not isinstance(parsed, dict):
                raise ValueError("fextract output must be a JSON object")
            return parsed, total_pt, total_ct, error_count, None
        except Exception as exc:
            last_error = exc
            error_count += 1
            retry_note = (
                "Your previous response was not valid JSON. "
                "Return only a valid JSON object with no markdown or extra text."
            )

    return None, total_pt, total_ct, error_count, f"fextract failed after retries: {last_error}"


FSCORING_SYSTEM = (
    "The Extracted Evidence provided in your context was generated by an upstream "
    "agent called fextract. For your understanding of the input data quality and context, "
    "here is the fextract agent profile:\n"
    "{The input data consists of evidence extracted by an upstream agent. "
    "You work by mapping specific text spans to specific rules, one at a time. "
    "Do not evaluate quality — only locate and classify. "
    "Follow the output instructions exactly.}"
)

FSCORING_OUTPUT_INSTRUCTIONS = (

    "Return json with grade and reasoning."
)
FSCORING_USER_TEMPLATE = """\
Question: {question}

Official Rubric:
{rubric}

Student Response:
{student_answer}

Extracted Evidence:
{extracted_json}

Output Instructions: {output_instructions}
"""


def build_fscoring_messages(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    retry_note: str = "",
) -> list[dict[str, str]]:
    user_content = FSCORING_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        extracted_json=json.dumps(extracted, ensure_ascii=False, indent=2),
        output_instructions=FSCORING_OUTPUT_INSTRUCTIONS,
    )
    if retry_note:
        user_content += f"\n\nIMPORTANT: {retry_note}"
    return [
        {"role": "system", "content": FSCORING_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def parse_fscoring_output(raw: str) -> dict[str, Any]:
    parsed = json.loads(raw)
    if not isinstance(parsed, dict):
        raise ValueError("fscoring output must be a JSON object")
    if "grade" not in parsed or "reasoning" not in parsed:
        raise ValueError("fscoring output must have 'grade' and 'reasoning'")
    grade = parsed["grade"]
    if not isinstance(grade, int) or not (1 <= grade <= 6):
        raise ValueError(f"grade must be an integer 1-6, got: {grade!r}")
    reasoning = parsed["reasoning"]
    if not isinstance(reasoning, str) or not reasoning.strip():
        raise ValueError("reasoning must be a non-empty string")
    return {"grade": grade, "reasoning": reasoning.strip()}


def run_fscoring(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any], int, int, int, str | None]:
    retry_note = ""
    last_error: Exception | None = None
    total_pt = 0
    total_ct = 0
    error_count = 0

    for attempt in range(2):
        messages = build_fscoring_messages(
            question, student_answer, rubric, extracted, retry_note
        )
        try:
            raw, pt, ct = call_azure_openai(
                messages,
                endpoint,
                api_key,
                deployment,
                api_version,
                max_tokens=DEFAULT_MAX_TOKENS_SCORE,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                break
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(
                f"  [fscoring {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s..."
            )
            time.sleep(wait_seconds)
            retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only valid JSON with exactly the keys 'grade' and 'reasoning'."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            return {"grade": -1, "reasoning": "Content filter triggered."}, total_pt, total_ct, error_count, None

        try:
            return parse_fscoring_output(raw), total_pt, total_ct, error_count, None
        except Exception as exc:
            last_error = exc
            error_count += 1
            retry_note = (
                "Your previous response was invalid. "
                "Return only valid JSON with exactly the keys 'grade' (int 1-6) and 'reasoning' (string)."
            )

    return (
        {"grade": -1, "reasoning": f"fscoring failed after retries: {last_error}"},
        total_pt,
        total_ct,
        error_count,
        f"fscoring failed after retries: {last_error}",
    )


_worker_local = threading.local()
_worker_counter = 0
_worker_counter_lock = threading.Lock()


def _assign_worker_id() -> int:
    if not hasattr(_worker_local, "id"):
        global _worker_counter
        with _worker_counter_lock:
            _worker_counter += 1
            _worker_local.id = _worker_counter
    return _worker_local.id


def grade_row(
    question: str,
    answer: str,
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
    worker_id: int,
) -> dict[str, Any]:
    t0 = time.perf_counter()

    extracted, a1_pt, a1_ct, a1_errors, a1_error = run_fextract(
        question, answer, rubric, endpoint, api_key, deployment, api_version
    )
    if extracted is None:
        wall = time.perf_counter() - t0
        if a1_error is None:
            return {
                "grade": -1,
                "reasoning": "Content filter triggered (extraction).",
                "extracted_evidence": "{}",
                "a1_prompt_tokens": a1_pt,
                "a1_completion_tokens": a1_ct,
                "a2_prompt_tokens": 0,
                "a2_completion_tokens": 0,
                "a1_errors": a1_errors,
                "a2_errors": 0,
                "wall_seconds": round(wall, 3),
                "worker_id": worker_id,
                "retryable_events": a1_errors,
                "status": "skip",
            }

        return {
            "grade": -1,
            "reasoning": a1_error,
            "extracted_evidence": "{}",
            "a1_prompt_tokens": a1_pt,
            "a1_completion_tokens": a1_ct,
            "a2_prompt_tokens": 0,
            "a2_completion_tokens": 0,
            "a1_errors": a1_errors,
            "a2_errors": 0,
            "wall_seconds": round(wall, 3),
            "worker_id": worker_id,
            "retryable_events": a1_errors,
            "status": "error",
        }

    scored, a2_pt, a2_ct, a2_errors, a2_error = run_fscoring(
        question, answer, rubric, extracted, endpoint, api_key, deployment, api_version
    )

    wall = time.perf_counter() - t0
    if a2_error is not None:
        status = "error"
    elif scored["grade"] == -1 and scored["reasoning"] == "Content filter triggered.":
        status = "skip"
    else:
        status = "ok"

    return {
        "grade": scored["grade"],
        "reasoning": scored["reasoning"],
        "extracted_evidence": json.dumps(extracted, ensure_ascii=False),
        "a1_prompt_tokens": a1_pt,
        "a1_completion_tokens": a1_ct,
        "a2_prompt_tokens": a2_pt,
        "a2_completion_tokens": a2_ct,
        "a1_errors": a1_errors,
        "a2_errors": a2_errors,
        "wall_seconds": round(wall, 3),
        "worker_id": worker_id,
        "retryable_events": a1_errors + a2_errors,
        "status": status,
    }


def _grade_row_wrapper(
    row_index: int,
    row: dict[str, str],
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[int, dict[str, str], dict[str, Any]]:
    worker_id = _assign_worker_id()
    question = (row.get("assignment") or "").strip()
    student_answer = (row.get("full_text") or "").strip()
    result = grade_row(
        question,
        student_answer,
        rubric,
        endpoint,
        api_key,
        deployment,
        api_version,
        worker_id,
    )
    return row_index, row, result


def append_summary_to_excel(
    xlsx_path: Path,
    run_start: datetime,
    run_end: datetime,
    duration_str: str,
    sequential_time: float,
    total_wall_seconds: float,
    speedup_factor: float,
    test_name: str,
    total_a1_prompt_tokens: int,
    total_a1_completion_tokens: int,
    total_a2_prompt_tokens: int,
    total_a2_completion_tokens: int,
    total_a1_errors: int,
    total_a2_errors: int,
    total_tokens: int,
    skipped_count: int,
    skipped_ids: list[str],
    start_workers: int,
    peak_workers: int,
    min_workers: int,
    retryable_events: int,
    error_rows: int,
) -> None:
    """Append one summary row to the Excel file; create with headers if it doesn't exist."""
    def normalize_headers(ws) -> list[str]:
        headers: list[str] = []
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            headers.append(str(value).strip() if value is not None else "")
        while headers and headers[-1] == "":
            headers.pop()
        return headers

    def align_sheet_headers(wb: Workbook, ws) -> None:
        existing_headers = normalize_headers(ws)
        if not existing_headers or existing_headers == SUMMARY_HEADERS:
            return

        aligned = wb.create_sheet(title="Grading Runs Aligned")
        for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
            aligned.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(
                bold=True, name="Arial"
            )

        for row_idx in range(2, ws.max_row + 1):
            row_map = {}
            for col_idx, header in enumerate(existing_headers, start=1):
                if header:
                    row_map[header] = ws.cell(row=row_idx, column=col_idx).value
            aligned.append([row_map.get(header, "") for header in SUMMARY_HEADERS])

        wb.remove(ws)
        aligned.title = "Grading Runs"

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        align_sheet_headers(wb, ws)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Grading Runs"

    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(
            bold=True, name="Arial"
        )

    col_widths = [25, 20, 20, 10, 16, 18, 14, 20, 22, 20, 22, 14, 14, 12, 14, 14, 12, 13, 40, 14, 14, 14, 16, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    new_row = [
        test_name,
        run_start.strftime("%Y-%m-%d %H:%M:%S"),
        run_end.strftime("%Y-%m-%d %H:%M:%S"),
        duration_str,
        f"{sequential_time:.3f}",
        f"{total_wall_seconds:.3f}",
        f"{speedup_factor:.2f}",
        total_a1_prompt_tokens,
        total_a1_completion_tokens,
        total_a2_prompt_tokens,
        total_a2_completion_tokens,
        total_a1_prompt_tokens + total_a2_prompt_tokens,
        total_a1_completion_tokens + total_a2_completion_tokens,
        total_tokens,
        total_a1_errors,
        total_a2_errors,
        total_a1_errors + total_a2_errors,
        skipped_count,
        ", ".join(skipped_ids) if skipped_ids else "",
        start_workers,
        peak_workers,
        min_workers,
        retryable_events,
        error_rows,
    ]
    ws.append(new_row)

    last_row = ws.max_row
    for col_idx in range(1, len(SUMMARY_HEADERS) + 1):
        ws.cell(row=last_row, column=col_idx).font = openpyxl.styles.Font(name="Arial")

    wb.save(xlsx_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--start-row",
        type=int,
        default=1,
        help="1-based row number in the input CSV to start from (rows before this are skipped).",
    )
    args = parser.parse_args()
    start_row = max(1, args.start_row)

    load_env_file(BASE_DIR / ".env")

    if not INPUT_CSV.exists():
        raise SystemExit(f"Input CSV not found: {INPUT_CSV}")

    endpoint = get_env("AZURE_OPENAI_ENDPOINT")
    api_key = get_env("AZURE_OPENAI_API_KEY")
    deployment = get_env("AZURE_OPENAI_DEPLOYMENT")
    api_version = get_env("AZURE_OPENAI_API_VERSION", DEFAULT_API_VERSION).strip()

    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as infile:
        reader = csv.DictReader(infile, delimiter=",")
        input_fieldnames = list(reader.fieldnames or [])
        all_rows = list(reader)

    if not all_rows:
        raise SystemExit("Input CSV is empty.")

    rows_to_skip = start_row - 1
    if rows_to_skip >= len(all_rows):
        raise SystemExit(
            f"--start-row {start_row} exceeds total row count ({len(all_rows)}). Nothing to do."
        )

    rows_to_process = all_rows[rows_to_skip:]
    for i, row in enumerate(rows_to_process, start=start_row):
        if not (row.get("assignment") or "").strip():
            raise SystemExit(f"Missing assignment in row {i}")
        if not (row.get("full_text") or "").strip():
            raise SystemExit(f"Missing full_text in row {i}")

    rubric_text = get_rubric()

    output_fieldnames = list(input_fieldnames)
    for extra_field in (
        "AI_grade",
        "AI_reasoning",
        "extracted_evidence",
        "a1_prompt_tokens",
        "a1_completion_tokens",
        "a2_prompt_tokens",
        "a2_completion_tokens",
        "prompt_tokens",
        "completion_tokens",
        "wall_seconds",
        "worker_id",
        "squential time",
        "total_wall_seconds",
        "speedup_factor",
        "adaptive_start_workers",
        "adaptive_peak_workers",
        "adaptive_min_workers",
        "adaptive_retryable_events",
        "adaptive_error_rows",
    ):
        if extra_field not in output_fieldnames:
            output_fieldnames.append(extra_field)

    run_start_time = datetime.now()
    start_log = OUTPUT_CSV.with_suffix(".starttime")
    start_log.write_text(run_start_time.strftime("%Y-%m-%d %H:%M:%S"), encoding="utf-8")

    start_workers = resolve_worker_count()
    growth_threshold = resolve_growth_threshold()
    target_workers = start_workers
    peak_target_workers = start_workers
    min_target_workers = start_workers
    clean_success_streak = 0

    total_a1_pt = 0
    total_a1_ct = 0
    total_a1_errors = 0
    total_a2_pt = 0
    total_a2_ct = 0
    total_a2_errors = 0
    total_sequential_time = 0.0
    adaptive_retryable_events = 0
    adaptive_error_rows = 0
    skipped_ids: list[str] = []

    write_mode = "a" if rows_to_skip > 0 and OUTPUT_CSV.exists() else "w"

    def new_row_state(row_index: int, row: dict[str, str]) -> dict[str, Any]:
        return {
            "row_index": row_index,
            "row": dict(row),
            "attempts": 0,
            "wall_seconds": 0.0,
            "a1_prompt_tokens": 0,
            "a1_completion_tokens": 0,
            "a2_prompt_tokens": 0,
            "a2_completion_tokens": 0,
            "a1_errors": 0,
            "a2_errors": 0,
            "worker_id": 0,
        }

    pending_rows: deque[dict[str, Any]] = deque(
        new_row_state(row_index, row)
        for row_index, row in enumerate(rows_to_process, start=start_row)
    )
    active_futures: dict[Any, dict[str, Any]] = {}
    results_buffer: dict[int, dict[str, str]] = {}
    completed_rows = 0
    next_to_write = start_row
    wall_start = time.perf_counter()

    def flush_buffer(writer: Any, outfile: Any) -> None:
        nonlocal next_to_write, completed_rows
        while next_to_write in results_buffer:
            writer.writerow(results_buffer.pop(next_to_write))
            outfile.flush()
            next_to_write += 1
            completed_rows += 1

    def adjust_workers(had_retryable_event: bool) -> None:
        nonlocal clean_success_streak, target_workers, peak_target_workers, min_target_workers
        if had_retryable_event:
            clean_success_streak = 0
            target_workers = max(200, target_workers - 1)
        else:
            clean_success_streak += 1
            if clean_success_streak >= growth_threshold and target_workers < start_workers:
                target_workers += 10
                clean_success_streak = 0
        peak_target_workers = max(peak_target_workers, target_workers)
        min_target_workers = min(min_target_workers, target_workers)

    run_end_time = run_start_time
    run_duration = run_end_time - run_start_time
    duration_str = "0:00:00"
    total_prompt = 0
    total_completion = 0
    total_tokens = 0
    total_errors = 0
    total_wall_seconds = 0.0
    speedup_factor = 0.0

    with OUTPUT_CSV.open(write_mode, encoding="utf-8", newline="") as outfile:
        writer = csv.DictWriter(
            outfile,
            fieldnames=output_fieldnames,
            delimiter=",",
            quoting=csv.QUOTE_MINIMAL,
        )
        if write_mode == "w":
            writer.writeheader()

        with ThreadPoolExecutor(max_workers=500) as executor:
            while pending_rows or active_futures:
                while pending_rows and len(active_futures) < target_workers:
                    state = pending_rows.popleft()
                    future = executor.submit(
                        _grade_row_wrapper,
                        state["row_index"],
                        state["row"],
                        rubric_text,
                        endpoint,
                        api_key,
                        deployment,
                        api_version,
                    )
                    active_futures[future] = state

                if not active_futures:
                    break

                done_futures, _ = wait(set(active_futures.keys()), return_when=FIRST_COMPLETED)

                for future in done_futures:
                    state = active_futures.pop(future)
                    row_index, _, result = future.result()

                    state["attempts"] += 1
                    state["wall_seconds"] += float(result["wall_seconds"])
                    state["a1_prompt_tokens"] += int(result["a1_prompt_tokens"])
                    state["a1_completion_tokens"] += int(result["a1_completion_tokens"])
                    state["a2_prompt_tokens"] += int(result["a2_prompt_tokens"])
                    state["a2_completion_tokens"] += int(result["a2_completion_tokens"])
                    state["a1_errors"] += int(result.get("a1_errors", 0))
                    state["a2_errors"] += int(result.get("a2_errors", 0))
                    state["worker_id"] = int(result["worker_id"])

                    retryable_events = int(result.get("retryable_events", 0))
                    if retryable_events > 0:
                        adaptive_retryable_events += retryable_events

                    if result["status"] == "error" and state["attempts"] < ROW_RETRY_LIMIT:
                        adjust_workers(True)
                        print(
                            f"Row {row_index} attempt {state['attempts']} failed; "
                            f"requeueing with target_workers={target_workers}."
                        )
                        pending_rows.appendleft(state)
                        continue

                    had_retryable_event = retryable_events > 0 or result["status"] == "error"
                    adjust_workers(had_retryable_event)

                    if result["status"] == "error":
                        adaptive_error_rows += 1

                    final_row = dict(state["row"])
                    final_row["AI_grade"] = str(result["grade"])
                    final_row["AI_reasoning"] = result["reasoning"]
                    final_row["extracted_evidence"] = result["extracted_evidence"]
                    final_row["a1_prompt_tokens"] = str(state["a1_prompt_tokens"])
                    final_row["a1_completion_tokens"] = str(state["a1_completion_tokens"])
                    final_row["a2_prompt_tokens"] = str(state["a2_prompt_tokens"])
                    final_row["a2_completion_tokens"] = str(state["a2_completion_tokens"])
                    final_row["prompt_tokens"] = str(state["a1_prompt_tokens"] + state["a2_prompt_tokens"])
                    final_row["completion_tokens"] = str(
                        state["a1_completion_tokens"] + state["a2_completion_tokens"]
                    )
                    final_row["wall_seconds"] = f"{state['wall_seconds']:.3f}"
                    final_row["worker_id"] = str(state["worker_id"])

                    results_buffer[row_index] = final_row
                    flush_buffer(writer, outfile)

                    total_a1_pt += state["a1_prompt_tokens"]
                    total_a1_ct += state["a1_completion_tokens"]
                    total_a1_errors += state["a1_errors"]
                    total_a2_pt += state["a2_prompt_tokens"]
                    total_a2_ct += state["a2_completion_tokens"]
                    total_a2_errors += state["a2_errors"]
                    total_sequential_time += state["wall_seconds"]

                    if result["status"] == "skip":
                        identifier = (
                            state["row"].get("essay_id")
                            or state["row"].get("identifier")
                            or f"row_{row_index}"
                        ).strip()
                        skipped_ids.append(identifier)

                    status = (
                        "ERROR (retry exhausted)"
                        if result["status"] == "error"
                        else "SKIPPED (content filter)"
                        if result["status"] == "skip"
                        else f"grade={result['grade']}"
                    )
                    print(
                        f"[{completed_rows:>4}/{len(rows_to_process)}] row={row_index:>4} "
                        f"worker={state['worker_id']} | {status} | "
                        f"target={target_workers} | wall={state['wall_seconds']:.1f}s | "
                        f"tokens A1 in={state['a1_prompt_tokens']} out={state['a1_completion_tokens']}  "
                        f"A2 in={state['a2_prompt_tokens']} out={state['a2_completion_tokens']}"
                    )

        if results_buffer:
            raise RuntimeError("Output order buffer was not fully flushed.")

        run_end_time = datetime.now()
        run_duration = run_end_time - run_start_time
        duration_str = str(run_duration).split(".")[0]
        total_prompt = total_a1_pt + total_a2_pt
        total_completion = total_a1_ct + total_a2_ct
        total_tokens = total_prompt + total_completion
        total_errors = total_a1_errors + total_a2_errors
        total_wall_seconds = time.perf_counter() - wall_start
        speedup_factor = (
            total_sequential_time / total_wall_seconds if total_wall_seconds > 0 else 0.0
        )

        append_summary_to_excel(
            xlsx_path=SUMMARY_XLSX,
            run_start=run_start_time,
            run_end=run_end_time,
            duration_str=duration_str,
            sequential_time=total_sequential_time,
            total_wall_seconds=total_wall_seconds,
            speedup_factor=speedup_factor,
            test_name=THIS_DIR.name,
            total_a1_prompt_tokens=total_a1_pt,
            total_a1_completion_tokens=total_a1_ct,
            total_a2_prompt_tokens=total_a2_pt,
            total_a2_completion_tokens=total_a2_ct,
            total_a1_errors=total_a1_errors,
            total_a2_errors=total_a2_errors,
            total_tokens=total_tokens,
            skipped_count=len(skipped_ids),
            skipped_ids=skipped_ids,
            start_workers=start_workers,
            peak_workers=peak_target_workers,
            min_workers=min_target_workers,
            retryable_events=adaptive_retryable_events,
            error_rows=adaptive_error_rows,
        )

        summary: dict[str, str] = {field: "" for field in output_fieldnames}
        summary["assignment"] = "=== RUN SUMMARY ==="
        summary["AI_reasoning"] = (
            f"start={run_start_time.strftime('%Y-%m-%d %H:%M:%S')} | "
            f"end={run_end_time.strftime('%Y-%m-%d %H:%M:%S')} | "
            f"duration={duration_str} | "
            f"sequential_time={total_sequential_time:.3f} | "
            f"total_wall_seconds={total_wall_seconds:.3f} | "
            f"speedup_factor={speedup_factor:.2f} | "
            f"workers={start_workers} | peak_workers={peak_target_workers} | min_workers={min_target_workers} | "
            f"retryable_events={adaptive_retryable_events} | error_rows={adaptive_error_rows} | "
            f"skipped={len(skipped_ids)} | "
            f"skipped_ids={', '.join(skipped_ids) if skipped_ids else 'none'}"
        )
        summary["a1_prompt_tokens"] = str(total_a1_pt)
        summary["a1_completion_tokens"] = str(total_a1_ct)
        summary["a2_prompt_tokens"] = str(total_a2_pt)
        summary["a2_completion_tokens"] = str(total_a2_ct)
        summary["prompt_tokens"] = str(total_prompt)
        summary["completion_tokens"] = str(total_completion)
        summary["wall_seconds"] = f"{total_wall_seconds:.3f}"
        summary["squential time"] = f"{total_sequential_time:.3f}"
        summary["total_wall_seconds"] = f"{total_wall_seconds:.3f}"
        summary["speedup_factor"] = f"{speedup_factor:.2f}"
        summary["adaptive_start_workers"] = str(start_workers)
        summary["adaptive_peak_workers"] = str(peak_target_workers)
        summary["adaptive_min_workers"] = str(min_target_workers)
        summary["adaptive_retryable_events"] = str(adaptive_retryable_events)
        summary["adaptive_error_rows"] = str(adaptive_error_rows)
        writer.writerow(summary)

    print("\n" + "=" * 60)
    print("AUTOSCORE RUN SUMMARY")
    print("=" * 60)
    print(f"  Start time              : {run_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  End time                : {run_end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Duration                : {duration_str}")
    print(f"  Sequential time         : {total_sequential_time:.3f}s")
    print(f"  Total wall seconds      : {total_wall_seconds:.3f}s")
    print(f"  Speedup factor          : {speedup_factor:.2f}x")
    print(f"  Start workers           : {start_workers}")
    print(f"  Peak workers            : {peak_target_workers}")
    print(f"  Min workers             : {min_target_workers}")
    print(f"  Retryable Azure events  : {adaptive_retryable_events}")
    print(f"  Error rows              : {adaptive_error_rows}")
    print(f"  Agent 1 prompt tokens   : {total_a1_pt:,}")
    print(f"  Agent 1 completion tok. : {total_a1_ct:,}")
    print(f"  Agent 2 prompt tokens   : {total_a2_pt:,}")
    print(f"  Agent 2 completion tok. : {total_a2_ct:,}")
    print(f"  -- Total prompt tokens  : {total_prompt:,}")
    print(f"  -- Total compl. tokens  : {total_completion:,}")
    print(f"  -- GRAND TOTAL tokens   : {total_tokens:,}")
    print("=" * 60)
    print(f"Output written to {OUTPUT_CSV}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
