from __future__ import annotations

import argparse
import csv
import json
import os
import socket
import threading
import time
import urllib.error
import urllib.request
from collections import deque
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
THIS_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / "asap2_total_master.csv"
OUTPUT_CSV = THIS_DIR / "MASGraded.csv"
SUMMARY_XLSX = BASE_DIR / "GradingSummary.xlsx"

RUBRIC_FILE = BASE_DIR / "Criteria.txt"

DEFAULT_API_VERSION = "2024-12-01-preview"
DEFAULT_TEMPERATURE = 0.0
DEFAULT_MAX_TOKENS_EXTRACT = 1000
DEFAULT_MAX_TOKENS_SCORE = 1000
DEFAULT_MAX_TOKENS_ADVOCATE = 800
DEFAULT_WORKERS = 150
ADAPTIVE_GROWTH_STREAK = 5
ROW_RETRY_LIMIT = 3
BASE_RETRY_DELAY_SECONDS = 2.0
MAX_RETRY_DELAY_SECONDS = 20.0

# ──────────────────────────────────────────────────────────────────────────────
# Devil's Advocate flag
#
# True  → the devil's advocate feedback is ALWAYS forwarded to fextract/fscoring
#          for reconsideration, regardless of whether the advocate recommends a
#          revision or not.
# False → the feedback is only forwarded when the advocate explicitly recommends
#          a revision (advocate JSON contains "recommend_revision": true).
# ──────────────────────────────────────────────────────────────────────────────
DEVILS_ADVOCATE_ALWAYS_SEND: bool = False

SUMMARY_HEADERS = [
    "Test Name",
    "Start Time",
    "End Time",
    "Duration",
    "Sequential Time",
    "Total Wall Seconds",
    "Speedup Factor",
    "Agent 1 Prompt Tokens",
    "Agent 1 Completion Tokens",
    "Agent 2 Prompt Tokens",
    "Agent 2 Completion Tokens",
    "Agent 3 Prompt Tokens",
    "Agent 3 Completion Tokens",
    "Agent 4 Prompt Tokens",
    "Agent 4 Completion Tokens",
    "Input Tokens",
    "Output Tokens",
    "Total Tokens",
    "Agent 1 Errors",
    "Agent 2 Errors",
    "Agent 3 Errors",
    "Agent 4 Errors",
    "Total Errors",
    "Skipped Rows",
    "Skipped Row IDs",
    "Start Workers",
    "Peak Workers",
    "Min Workers",
    "Retryable Events",
    "Error Rows",
    "A3 Revisions Recommended",
    "A4 Revisions Recommended",
    "A3 Revision IDs",
    "A4 Revision IDs",
]


# ── helpers ───────────────────────────────────────────────────────────────────

def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def get_env(name: str, default: str | None = None) -> str:
    value = os.getenv(name, default)
    if value is None or not value.strip():
        raise SystemExit(f"Missing required environment variable: {name}")
    return value.strip()


def get_rubric() -> str:
    if not RUBRIC_FILE.exists():
        raise SystemExit(f"Rubric file not found: {RUBRIC_FILE}")
    return load_text(RUBRIC_FILE)


def parse_retry_after(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def resolve_worker_count() -> int:
    raw = os.getenv("MAX_WORKERS", "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return DEFAULT_WORKERS


def resolve_growth_threshold() -> int:
    raw = os.getenv("ADAPTIVE_GROWTH_STREAK", "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return ADAPTIVE_GROWTH_STREAK


def retry_delay_seconds(attempt: int, retry_after: float | None = None) -> float:
    if retry_after is not None and retry_after > 0:
        return min(retry_after, MAX_RETRY_DELAY_SECONDS)
    delay = BASE_RETRY_DELAY_SECONDS * (2 ** attempt)
    return min(delay, MAX_RETRY_DELAY_SECONDS)


# ── Azure error classes ───────────────────────────────────────────────────────

class AzureRetryableError(RuntimeError):
    def __init__(self, message: str, *, kind: str, retry_after: float | None = None) -> None:
        super().__init__(message)
        self.kind = kind
        self.retry_after = retry_after


class AzureThrottleError(AzureRetryableError):
    def __init__(self, message: str, *, retry_after: float | None = None) -> None:
        super().__init__(message, kind="throttle", retry_after=retry_after)


class AzureTransientError(AzureRetryableError):
    def __init__(self, message: str, *, retry_after: float | None = None) -> None:
        super().__init__(message, kind="transient", retry_after=retry_after)


# ── Azure API call ────────────────────────────────────────────────────────────

def call_azure_openai(
    messages: list[dict[str, str]],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
    max_tokens: int,
) -> tuple[str | None, int, int]:
    """Returns (content, prompt_tokens, completion_tokens). Returns (None, 0, 0) on content-filter hit."""
    url = (
        f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
        f"/chat/completions?api-version={api_version}"
    )
    payload = {
        "messages": messages,
        "temperature": DEFAULT_TEMPERATURE,
        "max_tokens": max_tokens,
        "response_format": {"type": "json_object"},
    }
    request = urllib.request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "api-key": api_key},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            response_body = response.read().decode("utf-8")
    except TimeoutError as exc:
        raise AzureTransientError(
            f"Azure OpenAI request timed out after 120 seconds: {exc}"
        ) from exc
    except socket.timeout as exc:
        raise AzureTransientError(
            f"Azure OpenAI request timed out after 120 seconds: {exc}"
        ) from exc
    except urllib.error.URLError as exc:
        raise AzureTransientError(f"Azure OpenAI request failed: {exc}") from exc
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        if "content_filter" in error_body:
            print("  Content filter triggered. Skipping this row.")
            return None, 0, 0
        if exc.code == 429:
            raise AzureThrottleError(
                f"Azure OpenAI request throttled: {exc.code} {exc.reason}\n{error_body}",
                retry_after=parse_retry_after(exc.headers.get("Retry-After")),
            ) from exc
        if exc.code in {408, 500, 502, 503, 504}:
            raise AzureTransientError(
                f"Azure OpenAI request failed transiently: {exc.code} {exc.reason}\n{error_body}"
            ) from exc
        raise RuntimeError(
            f"Azure OpenAI request failed: {exc.code} {exc.reason}\n{error_body}"
        ) from exc

    response_json = json.loads(response_body)
    usage = response_json.get("usage", {})
    prompt_tokens = usage.get("prompt_tokens", 0)
    completion_tokens = usage.get("completion_tokens", 0)

    choices = response_json.get("choices", [])
    if not choices:
        raise RuntimeError(f"Azure OpenAI returned no choices: {response_body}")

    content = choices[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"Azure OpenAI returned empty content: {response_body}")

    return content, prompt_tokens, completion_tokens


# ── fextract ──────────────────────────────────────────────────────────────────

FEXTRACT_SYSTEM = ""

FEXTRACT_OUTPUT_INSTRUCTIONS = (
    "Perform a rule-by-rule check of the student response against every rubric criterion. "
    "Return a single valid JSON object and nothing else. No markdown, no extra text. "
    "For EVERY rubric criterion include a key with: "
    '"requirement_met" (boolean), '
    '"evidence" (exact text span from the student response, empty string if none), '
    'and "count" (integer, only when the rubric requires counting). '
    "Use concise descriptive key names derived from each criterion "
    '(e.g. "defines_concept", "lists_two_advantages"). '
    "Also include a top-level \"overall_notes\" string for any ambiguities."
)

FEXTRACT_USER_TEMPLATE = """\
Question: {question}

Rubric:
{rubric}

Student Response:
{student_answer}

Output Instructions: {output_instructions}
"""


def build_fextract_messages(
    question: str, student_answer: str, rubric: str, retry_note: str = ""
) -> list[dict[str, str]]:
    user_content = FEXTRACT_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        output_instructions=FEXTRACT_OUTPUT_INSTRUCTIONS,
    )
    if retry_note:
        user_content += f"\n\nIMPORTANT: {retry_note}"
    return [
        {"role": "system", "content": FEXTRACT_SYSTEM},
        {"role": "user", "content": user_content},
    ]


# ── fextract devil's advocate ─────────────────────────────────────────────────

FEXTRACT_ADVOCATE_SYSTEM = ""

FEXTRACT_ADVOCATE_OUTPUT_INSTRUCTIONS = (
    "You are a devil's advocate reviewing an evidence-extraction result. "
    "Critically examine every criterion in the Extracted Evidence JSON. "
    "Look for: incorrect requirement_met flags, missed evidence spans, "
    "over-generous or overly strict interpretations, and counting errors. "
    "Return ONLY a valid JSON object with exactly these keys:\n"
    '  "recommend_revision" (boolean): true if any criterion should be changed,\n'
    '  "critique" (object): one key per criterion that you challenge, each with\n'
    '    "issue" (string) and "suggested_fix" (string),\n'
    '  "overall_comment" (string): brief summary of the critique.\n'
    "No markdown, no extra text."
)

FEXTRACT_ADVOCATE_USER_TEMPLATE = """\
Question: {question}

Rubric:
{rubric}

Student Response:
{student_answer}

Extracted Evidence (to critique):
{extracted_json}

Output Instructions: {output_instructions}
"""


def build_fextract_advocate_messages(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
) -> list[dict[str, str]]:
    user_content = FEXTRACT_ADVOCATE_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        extracted_json=json.dumps(extracted, ensure_ascii=False, indent=2),
        output_instructions=FEXTRACT_ADVOCATE_OUTPUT_INSTRUCTIONS,
    )
    return [
        {"role": "system", "content": FEXTRACT_ADVOCATE_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def run_fextract_advocate(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any] | None, int, int, int]:
    """
    Returns (advocate_json, prompt_tokens, completion_tokens, error_count).
    Returns (None, 0, 0, 0) on content-filter hit; (None, pt, ct, 1) on parse failure.
    Raises AzureRetryableError on throttle/transient so the caller can handle it.
    """
    messages = build_fextract_advocate_messages(question, student_answer, rubric, extracted)
    raw, pt, ct = call_azure_openai(
        messages,
        endpoint,
        api_key,
        deployment,
        api_version,
        max_tokens=DEFAULT_MAX_TOKENS_ADVOCATE,
    )
    if raw is None:
        return None, pt, ct, 0

    try:
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            raise ValueError("fextract_advocate output must be a JSON object")
        if "recommend_revision" not in parsed:
            raise ValueError("fextract_advocate output must include 'recommend_revision'")
        return parsed, pt, ct, 0
    except Exception as exc:
        print(f"  [warn] fextract_advocate parse error (ignored): {exc}")
        return None, pt, ct, 1


FEXTRACT_REVISION_USER_TEMPLATE = """\
Question: {question}

Rubric:
{rubric}

Student Response:
{student_answer}

Your Previous Extracted Evidence:
{previous_extracted_json}

Devil's Advocate Critique:
{advocate_json}

Output Instructions: {output_instructions}

Re-examine your previous extraction in light of the critique above.
Accept corrections that are well-founded; keep your original judgement where the critique is wrong.
Return the final corrected JSON object (same schema as before).
"""


def build_fextract_revision_messages(
    question: str,
    student_answer: str,
    rubric: str,
    previous_extracted: dict[str, Any],
    advocate: dict[str, Any],
) -> list[dict[str, str]]:
    user_content = FEXTRACT_REVISION_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        previous_extracted_json=json.dumps(previous_extracted, ensure_ascii=False, indent=2),
        advocate_json=json.dumps(advocate, ensure_ascii=False, indent=2),
        output_instructions=FEXTRACT_OUTPUT_INSTRUCTIONS,
    )
    return [
        {"role": "system", "content": FEXTRACT_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def run_fextract(
    question: str,
    student_answer: str,
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None, int, int, int, int, int, int, int, str | None]:
    """
    Returns:
        (extracted, advocate_json,
         a1_pt, a1_ct, adv_pt, adv_ct,
         a1_error_count, adv_error_count,
         a3_revision_recommended,
         fatal_error_msg)

        extracted               – final extraction dict (or None on content-filter / fatal error)
        advocate_json           – advocate output dict (or None if not run / content-filter)
        a1_pt / a1_ct          – tokens for fextract initial + revision calls
        adv_pt / adv_ct        – tokens for the A3 advocate call
        a1_error_count         – parse/retry errors from fextract itself
        adv_error_count        – parse errors from the A3 advocate call
        a3_revision_recommended – 1 if the A3 advocate recommended a revision, 0 otherwise
        fatal_error_msg         – non-None string when the failure is a hard error (not content-filter)
    """
    retry_note = ""
    last_error: Exception | None = None
    total_pt, total_ct = 0, 0
    error_count = 0
    adv_pt, adv_ct = 0, 0
    adv_error_count = 0
    a3_revision_recommended = 0
    advocate_result: dict[str, Any] | None = None

    # ── initial extraction ────────────────────────────────────────────────────
    for attempt in range(2):
        messages = build_fextract_messages(question, student_answer, rubric, retry_note)
        try:
            raw, pt, ct = call_azure_openai(
                messages, endpoint, api_key, deployment, api_version,
                max_tokens=DEFAULT_MAX_TOKENS_EXTRACT,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                return None, None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count, \
                    0, f"fextract failed after retries: {exc}"
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(f"  [fextract {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s...")
            time.sleep(wait_seconds)
            retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only a valid JSON object with no markdown or extra text."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            return None, None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count, 0, None

        try:
            parsed = json.loads(raw)
            if not isinstance(parsed, dict):
                raise ValueError("fextract output must be a JSON object")
            initial_extracted = parsed
            break
        except Exception as exc:
            last_error = exc
            error_count += 1
            retry_note = (
                "Your previous response was not valid JSON. "
                "Return only a valid JSON object with no markdown or extra text."
            )
    else:
        return None, None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count, \
            0, f"fextract failed after retries: {last_error}"

    # ── devil's advocate (A3) ─────────────────────────────────────────────────
    try:
        advocate_result, adv_pt, adv_ct, adv_error_count = run_fextract_advocate(
            question, student_answer, rubric, initial_extracted,
            endpoint, api_key, deployment, api_version,
        )
    except AzureRetryableError as exc:
        # Advocate failure is non-fatal: log and continue without revision
        print(f"  [warn] fextract advocate {exc.kind} error (ignored): {exc}")
        advocate_result = None
        adv_error_count = 1
    except Exception as exc:
        print(f"  [warn] fextract advocate error (ignored): {exc}")
        advocate_result = None
        adv_error_count = 1

    should_revise = (
        advocate_result is not None
        and (
            DEVILS_ADVOCATE_ALWAYS_SEND
            or advocate_result.get("recommend_revision", False)
        )
    )

    if advocate_result is not None and advocate_result.get("recommend_revision", False):
        a3_revision_recommended = 1

    if not should_revise:
        return (
            initial_extracted, advocate_result,
            total_pt, total_ct, adv_pt, adv_ct,
            error_count, adv_error_count, a3_revision_recommended, None,
        )

    # ── revision call ─────────────────────────────────────────────────────────
    revision_retry_note = ""
    for attempt in range(2):
        rev_messages = build_fextract_revision_messages(
            question, student_answer, rubric, initial_extracted, advocate_result,
        )
        if revision_retry_note:
            rev_messages[-1]["content"] += f"\n\nIMPORTANT: {revision_retry_note}"
        try:
            raw, pt, ct = call_azure_openai(
                rev_messages, endpoint, api_key, deployment, api_version,
                max_tokens=DEFAULT_MAX_TOKENS_EXTRACT,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                # Revision transient failure: fall back to initial extraction
                print("  [warn] fextract revision transient error; using initial extraction.")
                return (
                    initial_extracted, advocate_result,
                    total_pt, total_ct, adv_pt, adv_ct,
                    error_count, adv_error_count, a3_revision_recommended, None,
                )
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(f"  [fextract revision {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s...")
            time.sleep(wait_seconds)
            revision_retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only a valid JSON object with no markdown or extra text."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            # content filter on revision – fall back to initial extraction
            return (
                initial_extracted, advocate_result,
                total_pt, total_ct, adv_pt, adv_ct,
                error_count, adv_error_count, a3_revision_recommended, None,
            )

        try:
            revised = json.loads(raw)
            if not isinstance(revised, dict):
                raise ValueError("fextract revision output must be a JSON object")
            return (
                revised, advocate_result,
                total_pt, total_ct, adv_pt, adv_ct,
                error_count, adv_error_count, a3_revision_recommended, None,
            )
        except Exception as exc:
            last_error = exc
            error_count += 1
            revision_retry_note = (
                "Your previous response was not valid JSON. "
                "Return only a valid JSON object with no markdown or extra text."
            )

    # revision parse failed twice – fall back to initial extraction
    print("  [warn] fextract revision failed; using initial extraction.")
    return (
        initial_extracted, advocate_result,
        total_pt, total_ct, adv_pt, adv_ct,
        error_count, adv_error_count, a3_revision_recommended, None,
    )


# ── fscoring ──────────────────────────────────────────────────────────────────

FSCORING_SYSTEM = ""

FSCORING_OUTPUT_INSTRUCTIONS = (
    "Use the boolean flags, evidence strings, and counts in the Extracted Evidence "
    "to determine the final score. "
    "If the Extracted Evidence contains inconsistencies, use the Student Response for verification. "
    "Resolve all ambiguities strictly in favour of the official rubric definitions. "
    "Return ONLY a valid JSON object with exactly two keys: "
    '"grade" (integer 1-6) and "reasoning" (one concise sentence). '
    "No markdown, no extra keys, no extra text."
)

FSCORING_USER_TEMPLATE = """\
Question: {question}

Official Rubric:
{rubric}

Student Response:
{student_answer}

Extracted Evidence:
{extracted_json}

Output Instructions: {output_instructions}
"""


def build_fscoring_messages(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    retry_note: str = "",
) -> list[dict[str, str]]:
    user_content = FSCORING_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        extracted_json=json.dumps(extracted, ensure_ascii=False, indent=2),
        output_instructions=FSCORING_OUTPUT_INSTRUCTIONS,
    )
    if retry_note:
        user_content += f"\n\nIMPORTANT: {retry_note}"
    return [
        {"role": "system", "content": FSCORING_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def parse_fscoring_output(raw: str) -> dict[str, Any]:
    parsed = json.loads(raw)
    if not isinstance(parsed, dict):
        raise ValueError("fscoring output must be a JSON object")
    if "grade" not in parsed or "reasoning" not in parsed:
        raise ValueError("fscoring output must have 'grade' and 'reasoning'")
    grade = parsed["grade"]
    if not isinstance(grade, int) or not (1 <= grade <= 6):
        raise ValueError(f"grade must be an integer 1-6, got: {grade!r}")
    reasoning = parsed["reasoning"]
    if not isinstance(reasoning, str) or not reasoning.strip():
        raise ValueError("reasoning must be a non-empty string")
    return {"grade": grade, "reasoning": reasoning.strip()}


# ── fscoring devil's advocate ─────────────────────────────────────────────────

FSCORING_ADVOCATE_SYSTEM = ""

FSCORING_ADVOCATE_OUTPUT_INSTRUCTIONS = (
    "You are a devil's advocate reviewing a grade assignment. "
    "Scrutinise whether the grade and reasoning faithfully follow the rubric. "
    "Look for: grade too high or too low, reasoning that misquotes the rubric, "
    "ignored evidence, or incorrect boundary decisions. "
    "Return ONLY a valid JSON object with exactly these keys:\n"
    '  "recommend_revision" (boolean): true if the grade or reasoning should change,\n'
    '  "suggested_grade" (integer 1-6 or null if no change),\n'
    '  "critique" (string): concise explanation of any issues found,\n'
    '  "overall_comment" (string): brief summary.\n'
    "No markdown, no extra text."
)

FSCORING_ADVOCATE_USER_TEMPLATE = """\
Question: {question}

Official Rubric:
{rubric}

Student Response:
{student_answer}

Extracted Evidence:
{extracted_json}

Grade & Reasoning (to critique):
{scored_json}

Output Instructions: {output_instructions}
"""


def build_fscoring_advocate_messages(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    scored: dict[str, Any],
) -> list[dict[str, str]]:
    user_content = FSCORING_ADVOCATE_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        extracted_json=json.dumps(extracted, ensure_ascii=False, indent=2),
        scored_json=json.dumps(scored, ensure_ascii=False, indent=2),
        output_instructions=FSCORING_ADVOCATE_OUTPUT_INSTRUCTIONS,
    )
    return [
        {"role": "system", "content": FSCORING_ADVOCATE_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def run_fscoring_advocate(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    scored: dict[str, Any],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any] | None, int, int, int]:
    """
    Returns (advocate_json, prompt_tokens, completion_tokens, error_count).
    Returns (None, 0, 0, 0) on content-filter hit; (None, pt, ct, 1) on parse failure.
    Raises AzureRetryableError on throttle/transient so the caller can handle it.
    """
    messages = build_fscoring_advocate_messages(
        question, student_answer, rubric, extracted, scored
    )
    raw, pt, ct = call_azure_openai(
        messages, endpoint, api_key, deployment, api_version,
        max_tokens=DEFAULT_MAX_TOKENS_ADVOCATE,
    )
    if raw is None:
        return None, pt, ct, 0

    try:
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            raise ValueError("fscoring_advocate output must be a JSON object")
        if "recommend_revision" not in parsed:
            raise ValueError("fscoring_advocate output must include 'recommend_revision'")
        return parsed, pt, ct, 0
    except Exception as exc:
        print(f"  [warn] fscoring_advocate parse error (ignored): {exc}")
        return None, pt, ct, 1


FSCORING_REVISION_USER_TEMPLATE = """\
Question: {question}

Official Rubric:
{rubric}

Student Response:
{student_answer}

Extracted Evidence:
{extracted_json}

Your Previous Grade & Reasoning:
{previous_scored_json}

Devil's Advocate Critique:
{advocate_json}

Output Instructions: {output_instructions}

Re-examine your previous grade in light of the critique above.
Accept corrections that are well-founded; keep your original judgement where the critique is wrong.
Return the final corrected JSON object with exactly the keys "grade" (int 1-6) and "reasoning" (string).
"""


def build_fscoring_revision_messages(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    previous_scored: dict[str, Any],
    advocate: dict[str, Any],
) -> list[dict[str, str]]:
    user_content = FSCORING_REVISION_USER_TEMPLATE.format(
        question=question,
        rubric=rubric,
        student_answer=student_answer,
        extracted_json=json.dumps(extracted, ensure_ascii=False, indent=2),
        previous_scored_json=json.dumps(previous_scored, ensure_ascii=False, indent=2),
        advocate_json=json.dumps(advocate, ensure_ascii=False, indent=2),
        output_instructions=FSCORING_OUTPUT_INSTRUCTIONS,
    )
    return [
        {"role": "system", "content": FSCORING_SYSTEM},
        {"role": "user", "content": user_content},
    ]


def run_fscoring(
    question: str,
    student_answer: str,
    rubric: str,
    extracted: dict[str, Any],
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[dict[str, Any], dict[str, Any] | None, int, int, int, int, int, int, int, str | None]:
    """
    Returns:
        (scored, advocate_json,
         a2_pt, a2_ct, adv_pt, adv_ct,
         a2_error_count, adv_error_count,
         a4_revision_recommended,
         fatal_error_msg)

        scored                  – final grade dict
        advocate_json           – advocate output dict (or None if not run / content-filter)
        a2_pt / a2_ct          – tokens for fscoring initial + revision calls
        adv_pt / adv_ct        – tokens for the A4 advocate call
        a2_error_count         – parse/retry errors from fscoring itself
        adv_error_count        – parse errors from the A4 advocate call
        a4_revision_recommended – 1 if the A4 advocate recommended a revision, 0 otherwise
        fatal_error_msg         – non-None string when the failure is a hard error
    """
    retry_note = ""
    last_error: Exception | None = None
    total_pt, total_ct = 0, 0
    error_count = 0
    adv_pt, adv_ct = 0, 0
    adv_error_count = 0
    a4_revision_recommended = 0
    advocate_result: dict[str, Any] | None = None

    # ── initial scoring ───────────────────────────────────────────────────────
    for attempt in range(2):
        messages = build_fscoring_messages(
            question, student_answer, rubric, extracted, retry_note
        )
        try:
            raw, pt, ct = call_azure_openai(
                messages, endpoint, api_key, deployment, api_version,
                max_tokens=DEFAULT_MAX_TOKENS_SCORE,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                return (
                    {"grade": -1, "reasoning": f"fscoring failed after retries: {exc}"},
                    None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count,
                    0, f"fscoring failed after retries: {exc}",
                )
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(f"  [fscoring {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s...")
            time.sleep(wait_seconds)
            retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only valid JSON with exactly the keys 'grade' and 'reasoning'."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            return (
                {"grade": -1, "reasoning": "Content filter triggered."},
                None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count, 0, None,
            )

        try:
            initial_scored = parse_fscoring_output(raw)
            break
        except Exception as exc:
            last_error = exc
            error_count += 1
            retry_note = (
                "Your previous response was invalid. "
                "Return only valid JSON with exactly the keys 'grade' (int 1-6) and 'reasoning' (string)."
            )
    else:
        return (
            {"grade": -1, "reasoning": f"fscoring failed after retries: {last_error}"},
            None, total_pt, total_ct, adv_pt, adv_ct, error_count, adv_error_count,
            0, f"fscoring failed after retries: {last_error}",
        )

    # ── devil's advocate (A4) ─────────────────────────────────────────────────
    try:
        advocate_result, adv_pt, adv_ct, adv_error_count = run_fscoring_advocate(
            question, student_answer, rubric, extracted, initial_scored,
            endpoint, api_key, deployment, api_version,
        )
    except AzureRetryableError as exc:
        print(f"  [warn] fscoring advocate {exc.kind} error (ignored): {exc}")
        advocate_result = None
        adv_error_count = 1
    except Exception as exc:
        print(f"  [warn] fscoring advocate error (ignored): {exc}")
        advocate_result = None
        adv_error_count = 1

    should_revise = (
        advocate_result is not None
        and (
            DEVILS_ADVOCATE_ALWAYS_SEND
            or advocate_result.get("recommend_revision", False)
        )
    )

    if advocate_result is not None and advocate_result.get("recommend_revision", False):
        a4_revision_recommended = 1

    if not should_revise:
        return (
            initial_scored, advocate_result,
            total_pt, total_ct, adv_pt, adv_ct,
            error_count, adv_error_count, a4_revision_recommended, None,
        )

    # ── revision call ─────────────────────────────────────────────────────────
    revision_retry_note = ""
    for attempt in range(2):
        rev_messages = build_fscoring_revision_messages(
            question, student_answer, rubric, extracted, initial_scored, advocate_result,
        )
        if revision_retry_note:
            rev_messages[-1]["content"] += f"\n\nIMPORTANT: {revision_retry_note}"
        try:
            raw, pt, ct = call_azure_openai(
                rev_messages, endpoint, api_key, deployment, api_version,
                max_tokens=DEFAULT_MAX_TOKENS_SCORE,
            )
        except AzureRetryableError as exc:
            error_count += 1
            last_error = exc
            if attempt >= 1:
                print("  [warn] fscoring revision transient error; using initial score.")
                return (
                    initial_scored, advocate_result,
                    total_pt, total_ct, adv_pt, adv_ct,
                    error_count, adv_error_count, a4_revision_recommended, None,
                )
            wait_seconds = retry_delay_seconds(attempt, exc.retry_after)
            print(f"  [fscoring revision {exc.kind}] attempt {attempt + 1}/2, retrying in {wait_seconds:.1f}s...")
            time.sleep(wait_seconds)
            revision_retry_note = (
                f"The previous attempt failed with a transient Azure {exc.kind} error. "
                "Return only valid JSON with exactly the keys 'grade' and 'reasoning'."
            )
            continue

        total_pt += pt
        total_ct += ct

        if raw is None:
            return (
                initial_scored, advocate_result,
                total_pt, total_ct, adv_pt, adv_ct,
                error_count, adv_error_count, a4_revision_recommended, None,
            )

        try:
            revised = parse_fscoring_output(raw)
            return (
                revised, advocate_result,
                total_pt, total_ct, adv_pt, adv_ct,
                error_count, adv_error_count, a4_revision_recommended, None,
            )
        except Exception as exc:
            last_error = exc
            error_count += 1
            revision_retry_note = (
                "Your previous response was invalid. "
                "Return only valid JSON with exactly the keys 'grade' (int 1-6) and 'reasoning' (string)."
            )

    print("  [warn] fscoring revision failed; using initial score.")
    return (
        initial_scored, advocate_result,
        total_pt, total_ct, adv_pt, adv_ct,
        error_count, adv_error_count, a4_revision_recommended, None,
    )


# ── grade_row ─────────────────────────────────────────────────────────────────

_worker_local = threading.local()
_worker_counter = 0
_worker_counter_lock = threading.Lock()


def _assign_worker_id() -> int:
    if not hasattr(_worker_local, "id"):
        global _worker_counter
        with _worker_counter_lock:
            _worker_counter += 1
            _worker_local.id = _worker_counter
    return _worker_local.id


def grade_row(
    question: str,
    answer: str,
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
    worker_id: int,
) -> dict[str, Any]:
    """
    Returns a result dict with keys:
        grade, reasoning, extracted_evidence,
        fextract_advocate_json, fscoring_advocate_json,
        a1_prompt_tokens, a1_completion_tokens,
        a2_prompt_tokens, a2_completion_tokens,
        a3_prompt_tokens, a3_completion_tokens,
        a4_prompt_tokens, a4_completion_tokens,
        a1_errors, a2_errors, a3_errors, a4_errors,
        a3_revision_recommended, a4_revision_recommended,
        wall_seconds, worker_id, retryable_events, status
    """
    t0 = time.perf_counter()

    (
        extracted, fextract_advocate,
        a1_pt, a1_ct, a3_pt, a3_ct,
        a1_errors, a3_errors,
        a3_revision_recommended,
        a1_fatal,
    ) = run_fextract(question, answer, rubric, endpoint, api_key, deployment, api_version)

    wall = time.perf_counter() - t0

    if extracted is None:
        status = "error" if a1_fatal else "skip"
        return {
            "grade": -1,
            "reasoning": a1_fatal or "Content filter triggered (extraction).",
            "extracted_evidence": "{}",
            "fextract_advocate_json": "{}",
            "fscoring_advocate_json": "{}",
            "a1_prompt_tokens": a1_pt,
            "a1_completion_tokens": a1_ct,
            "a2_prompt_tokens": 0,
            "a2_completion_tokens": 0,
            "a3_prompt_tokens": a3_pt,
            "a3_completion_tokens": a3_ct,
            "a4_prompt_tokens": 0,
            "a4_completion_tokens": 0,
            "a1_errors": a1_errors,
            "a2_errors": 0,
            "a3_errors": a3_errors,
            "a4_errors": 0,
            "a3_revision_recommended": a3_revision_recommended,
            "a4_revision_recommended": 0,
            "wall_seconds": round(wall, 3),
            "worker_id": worker_id,
            "retryable_events": a1_errors + a3_errors,
            "status": status,
        }

    (
        scored, fscoring_advocate,
        a2_pt, a2_ct, a4_pt, a4_ct,
        a2_errors, a4_errors,
        a4_revision_recommended,
        a2_fatal,
    ) = run_fscoring(question, answer, rubric, extracted, endpoint, api_key, deployment, api_version)

    wall = time.perf_counter() - t0

    if a2_fatal:
        status = "error"
    elif scored["grade"] == -1 and scored["reasoning"] == "Content filter triggered.":
        status = "skip"
    else:
        status = "ok"

    return {
        "grade": scored["grade"],
        "reasoning": scored["reasoning"],
        "extracted_evidence": json.dumps(extracted, ensure_ascii=False),
        "fextract_advocate_json": json.dumps(fextract_advocate, ensure_ascii=False) if fextract_advocate else "{}",
        "fscoring_advocate_json": json.dumps(fscoring_advocate, ensure_ascii=False) if fscoring_advocate else "{}",
        "a1_prompt_tokens": a1_pt,
        "a1_completion_tokens": a1_ct,
        "a2_prompt_tokens": a2_pt,
        "a2_completion_tokens": a2_ct,
        "a3_prompt_tokens": a3_pt,
        "a3_completion_tokens": a3_ct,
        "a4_prompt_tokens": a4_pt,
        "a4_completion_tokens": a4_ct,
        "a1_errors": a1_errors,
        "a2_errors": a2_errors,
        "a3_errors": a3_errors,
        "a4_errors": a4_errors,
        "a3_revision_recommended": a3_revision_recommended,
        "a4_revision_recommended": a4_revision_recommended,
        "wall_seconds": round(wall, 3),
        "worker_id": worker_id,
        "retryable_events": a1_errors + a2_errors + a3_errors + a4_errors,
        "status": status,
    }


def _grade_row_wrapper(
    row_index: int,
    row: dict[str, str],
    rubric: str,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
) -> tuple[int, dict[str, str], dict[str, Any]]:
    worker_id = _assign_worker_id()
    question = (row.get("assignment") or "").strip()
    student_answer = (row.get("full_text") or "").strip()
    result = grade_row(
        question,
        student_answer,
        rubric,
        endpoint,
        api_key,
        deployment,
        api_version,
        worker_id,
    )
    return row_index, row, result


# ── Excel summary ─────────────────────────────────────────────────────────────

def append_summary_to_excel(
    xlsx_path: Path,
    run_start: datetime,
    run_end: datetime,
    duration_str: str,
    sequential_time: float,
    total_wall_seconds: float,
    speedup_factor: float,
    test_name: str,
    total_a1_prompt_tokens: int,
    total_a1_completion_tokens: int,
    total_a2_prompt_tokens: int,
    total_a2_completion_tokens: int,
    total_a3_prompt_tokens: int,
    total_a3_completion_tokens: int,
    total_a4_prompt_tokens: int,
    total_a4_completion_tokens: int,
    total_a1_errors: int,
    total_a2_errors: int,
    total_a3_errors: int,
    total_a4_errors: int,
    total_tokens: int,
    skipped_count: int,
    skipped_ids: list[str],
    start_workers: int,
    peak_workers: int,
    min_workers: int,
    retryable_events: int,
    error_rows: int,
    total_a3_revisions_recommended: int,
    total_a4_revisions_recommended: int,
    a3_revision_ids: list[str],
    a4_revision_ids: list[str],
) -> None:
    """Append one summary row to the Excel file; create with headers if it doesn't exist."""

    def normalize_headers(ws) -> list[str]:
        headers: list[str] = []
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            headers.append(str(value).strip() if value is not None else "")
        while headers and headers[-1] == "":
            headers.pop()
        return headers

    def align_sheet_headers(wb: Workbook, ws) -> None:
        existing_headers = normalize_headers(ws)
        if not existing_headers or existing_headers == SUMMARY_HEADERS:
            return
        aligned = wb.create_sheet(title="Grading Runs Aligned")
        for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
            aligned.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(
                bold=True, name="Arial"
            )
        for row_idx in range(2, ws.max_row + 1):
            row_map = {}
            for col_idx, header in enumerate(existing_headers, start=1):
                if header:
                    row_map[header] = ws.cell(row=row_idx, column=col_idx).value
            aligned.append([row_map.get(header, "") for header in SUMMARY_HEADERS])
        wb.remove(ws)
        aligned.title = "Grading Runs"

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        align_sheet_headers(wb, ws)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Grading Runs"

    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(
            bold=True, name="Arial"
        )

    col_widths = [
        25, 20, 20, 10, 16, 18, 14,
        20, 22, 20, 22, 20, 22, 20, 22,
        14, 14, 12,
        14, 14, 14, 14, 12,
        13, 40,
        14, 14, 14, 16, 12, 16, 16, 40, 40,
    ]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    total_input = (
        total_a1_prompt_tokens + total_a2_prompt_tokens
        + total_a3_prompt_tokens + total_a4_prompt_tokens
    )
    total_output = (
        total_a1_completion_tokens + total_a2_completion_tokens
        + total_a3_completion_tokens + total_a4_completion_tokens
    )

    new_row = [
        test_name,
        run_start.strftime("%Y-%m-%d %H:%M:%S"),
        run_end.strftime("%Y-%m-%d %H:%M:%S"),
        duration_str,
        f"{sequential_time:.3f}",
        f"{total_wall_seconds:.3f}",
        f"{speedup_factor:.2f}",
        total_a1_prompt_tokens,
        total_a1_completion_tokens,
        total_a2_prompt_tokens,
        total_a2_completion_tokens,
        total_a3_prompt_tokens,
        total_a3_completion_tokens,
        total_a4_prompt_tokens,
        total_a4_completion_tokens,
        total_input,
        total_output,
        total_tokens,
        total_a1_errors,
        total_a2_errors,
        total_a3_errors,
        total_a4_errors,
        total_a1_errors + total_a2_errors + total_a3_errors + total_a4_errors,
        skipped_count,
        ", ".join(skipped_ids) if skipped_ids else "",
        start_workers,
        peak_workers,
        min_workers,
        retryable_events,
        error_rows,
        total_a3_revisions_recommended,
        total_a4_revisions_recommended,
        ", ".join(a3_revision_ids) if a3_revision_ids else "",
        ", ".join(a4_revision_ids) if a4_revision_ids else "",
    ]
    ws.append(new_row)

    last_row = ws.max_row
    for col_idx in range(1, len(SUMMARY_HEADERS) + 1):
        ws.cell(row=last_row, column=col_idx).font = openpyxl.styles.Font(name="Arial")

    wb.save(xlsx_path)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--start-row",
        type=int,
        default=1,
        help="1-based row number in the input CSV to start from (rows before this are skipped).",
    )
    args = parser.parse_args()
    start_row = max(1, args.start_row)

    load_env_file(BASE_DIR / ".env")

    if not INPUT_CSV.exists():
        raise SystemExit(f"Input CSV not found: {INPUT_CSV}")

    endpoint = get_env("AZURE_OPENAI_ENDPOINT")
    api_key = get_env("AZURE_OPENAI_API_KEY")
    deployment = get_env("AZURE_OPENAI_DEPLOYMENT")
    api_version = get_env("AZURE_OPENAI_API_VERSION", DEFAULT_API_VERSION).strip()

    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as infile:
        reader = csv.DictReader(infile, delimiter=",")
        input_fieldnames = list(reader.fieldnames or [])
        all_rows = list(reader)

    if not all_rows:
        raise SystemExit("Input CSV is empty.")

    rows_to_skip = start_row - 1
    if rows_to_skip >= len(all_rows):
        raise SystemExit(
            f"--start-row {start_row} exceeds total row count ({len(all_rows)}). Nothing to do."
        )

    rows_to_process = all_rows[rows_to_skip:]
    for i, row in enumerate(rows_to_process, start=start_row):
        if not (row.get("assignment") or "").strip():
            raise SystemExit(f"Missing assignment in row {i}")
        if not (row.get("full_text") or "").strip():
            raise SystemExit(f"Missing full_text in row {i}")

    rubric_text = get_rubric()

    output_fieldnames = list(input_fieldnames)
    for extra_field in (
        "AI_grade",
        "AI_reasoning",
        "extracted_evidence",
        "fextract_advocate_json",
        "fscoring_advocate_json",
        "a1_prompt_tokens",
        "a1_completion_tokens",
        "a2_prompt_tokens",
        "a2_completion_tokens",
        "a3_prompt_tokens",
        "a3_completion_tokens",
        "a4_prompt_tokens",
        "a4_completion_tokens",
        "a1_errors",
        "a2_errors",
        "a3_errors",
        "a4_errors",
        "a3_revision_recommended",
        "a4_revision_recommended",
        "prompt_tokens",
        "completion_tokens",
        "wall_seconds",
        "worker_id",
        "sequential_time",
        "total_wall_seconds",
        "speedup_factor",
        "adaptive_start_workers",
        "adaptive_peak_workers",
        "adaptive_min_workers",
        "adaptive_retryable_events",
        "adaptive_error_rows",
        "a3_revision_recommended_ids",
        "a4_revision_recommended_ids",
    ):
        if extra_field not in output_fieldnames:
            output_fieldnames.append(extra_field)

    run_start_time = datetime.now()
    start_log = OUTPUT_CSV.with_suffix(".starttime")
    start_log.write_text(run_start_time.strftime("%Y-%m-%d %H:%M:%S"), encoding="utf-8")

    start_workers = resolve_worker_count()
    growth_threshold = resolve_growth_threshold()
    target_workers = start_workers
    peak_target_workers = start_workers
    min_target_workers = start_workers
    clean_success_streak = 0

    total_a1_pt = 0
    total_a1_ct = 0
    total_a1_errors = 0
    total_a2_pt = 0
    total_a2_ct = 0
    total_a2_errors = 0
    total_a3_pt = 0
    total_a3_ct = 0
    total_a3_errors = 0
    total_a4_pt = 0
    total_a4_ct = 0
    total_a4_errors = 0
    total_a3_revisions_recommended = 0
    total_a4_revisions_recommended = 0
    total_sequential_time = 0.0
    adaptive_retryable_events = 0
    adaptive_error_rows = 0
    skipped_ids: list[str] = []
    a3_revision_ids: list[str] = []
    a4_revision_ids: list[str] = []

    write_mode = "a" if rows_to_skip > 0 and OUTPUT_CSV.exists() else "w"

    def new_row_state(row_index: int, row: dict[str, str]) -> dict[str, Any]:
        return {
            "row_index": row_index,
            "row": dict(row),
            "attempts": 0,
            "wall_seconds": 0.0,
            "a1_prompt_tokens": 0,
            "a1_completion_tokens": 0,
            "a2_prompt_tokens": 0,
            "a2_completion_tokens": 0,
            "a3_prompt_tokens": 0,
            "a3_completion_tokens": 0,
            "a4_prompt_tokens": 0,
            "a4_completion_tokens": 0,
            "a1_errors": 0,
            "a2_errors": 0,
            "a3_errors": 0,
            "a4_errors": 0,
            "a3_revision_recommended": 0,
            "a4_revision_recommended": 0,
            "worker_id": 0,
        }

    pending_rows: deque[dict[str, Any]] = deque(
        new_row_state(row_index, row)
        for row_index, row in enumerate(rows_to_process, start=start_row)
    )
    active_futures: dict[Any, dict[str, Any]] = {}
    results_buffer: dict[int, dict[str, str]] = {}
    completed_rows = 0
    next_to_write = start_row
    wall_start = time.perf_counter()

    def flush_buffer(writer: Any, outfile: Any) -> None:
        nonlocal next_to_write, completed_rows
        while next_to_write in results_buffer:
            writer.writerow(results_buffer.pop(next_to_write))
            outfile.flush()
            next_to_write += 1
            completed_rows += 1

    def adjust_workers(had_retryable_event: bool) -> None:
        nonlocal clean_success_streak, target_workers, peak_target_workers, min_target_workers
        if had_retryable_event:
            clean_success_streak = 0
            target_workers = max(100, target_workers - 1)
        else:
            clean_success_streak += 1
            if clean_success_streak >= growth_threshold and target_workers < start_workers:
                target_workers += 10
                clean_success_streak = 0
        peak_target_workers = max(peak_target_workers, target_workers)
        min_target_workers = min(min_target_workers, target_workers)

    # initialise summary variables so they're always defined
    run_end_time = run_start_time
    duration_str = "0:00:00"
    total_prompt = 0
    total_completion = 0
    total_tokens = 0
    total_wall_seconds = 0.0
    speedup_factor = 0.0

    with OUTPUT_CSV.open(write_mode, encoding="utf-8", newline="") as outfile:
        writer = csv.DictWriter(
            outfile,
            fieldnames=output_fieldnames,
            delimiter=",",
            quoting=csv.QUOTE_MINIMAL,
        )
        if write_mode == "w":
            writer.writeheader()

        with ThreadPoolExecutor(max_workers=500) as executor:
            while pending_rows or active_futures:
                # fill up to target_workers
                while pending_rows and len(active_futures) < target_workers:
                    state = pending_rows.popleft()
                    future = executor.submit(
                        _grade_row_wrapper,
                        state["row_index"],
                        state["row"],
                        rubric_text,
                        endpoint,
                        api_key,
                        deployment,
                        api_version,
                    )
                    active_futures[future] = state

                if not active_futures:
                    break

                done_futures, _ = wait(set(active_futures.keys()), return_when=FIRST_COMPLETED)

                for future in done_futures:
                    state = active_futures.pop(future)
                    row_index, _, result = future.result()

                    state["attempts"] += 1
                    state["wall_seconds"] += float(result["wall_seconds"])
                    state["a1_prompt_tokens"] += int(result["a1_prompt_tokens"])
                    state["a1_completion_tokens"] += int(result["a1_completion_tokens"])
                    state["a2_prompt_tokens"] += int(result["a2_prompt_tokens"])
                    state["a2_completion_tokens"] += int(result["a2_completion_tokens"])
                    state["a3_prompt_tokens"] += int(result["a3_prompt_tokens"])
                    state["a3_completion_tokens"] += int(result["a3_completion_tokens"])
                    state["a4_prompt_tokens"] += int(result["a4_prompt_tokens"])
                    state["a4_completion_tokens"] += int(result["a4_completion_tokens"])
                    state["a1_errors"] += int(result.get("a1_errors", 0))
                    state["a2_errors"] += int(result.get("a2_errors", 0))
                    state["a3_errors"] += int(result.get("a3_errors", 0))
                    state["a4_errors"] += int(result.get("a4_errors", 0))
                    state["a3_revision_recommended"] += int(result.get("a3_revision_recommended", 0))
                    state["a4_revision_recommended"] += int(result.get("a4_revision_recommended", 0))
                    state["worker_id"] = int(result["worker_id"])

                    retryable_events = int(result.get("retryable_events", 0))
                    if retryable_events > 0:
                        adaptive_retryable_events += retryable_events

                    # row-level retry on hard errors
                    if result["status"] == "error" and state["attempts"] < ROW_RETRY_LIMIT:
                        adjust_workers(True)
                        print(
                            f"Row {row_index} attempt {state['attempts']} failed; "
                            f"requeueing with target_workers={target_workers}."
                        )
                        pending_rows.appendleft(state)
                        continue

                    had_retryable_event = retryable_events > 0 or result["status"] == "error"
                    adjust_workers(had_retryable_event)

                    if result["status"] == "error":
                        adaptive_error_rows += 1

                    final_row = dict(state["row"])
                    final_row["AI_grade"] = str(result["grade"])
                    final_row["AI_reasoning"] = result["reasoning"]
                    final_row["extracted_evidence"] = result["extracted_evidence"]
                    final_row["fextract_advocate_json"] = result.get("fextract_advocate_json", "{}")
                    final_row["fscoring_advocate_json"] = result.get("fscoring_advocate_json", "{}")
                    final_row["a1_prompt_tokens"] = str(state["a1_prompt_tokens"])
                    final_row["a1_completion_tokens"] = str(state["a1_completion_tokens"])
                    final_row["a2_prompt_tokens"] = str(state["a2_prompt_tokens"])
                    final_row["a2_completion_tokens"] = str(state["a2_completion_tokens"])
                    final_row["a3_prompt_tokens"] = str(state["a3_prompt_tokens"])
                    final_row["a3_completion_tokens"] = str(state["a3_completion_tokens"])
                    final_row["a4_prompt_tokens"] = str(state["a4_prompt_tokens"])
                    final_row["a4_completion_tokens"] = str(state["a4_completion_tokens"])
                    final_row["a1_errors"] = str(state["a1_errors"])
                    final_row["a2_errors"] = str(state["a2_errors"])
                    final_row["a3_errors"] = str(state["a3_errors"])
                    final_row["a4_errors"] = str(state["a4_errors"])
                    final_row["a3_revision_recommended"] = str(state["a3_revision_recommended"])
                    final_row["a4_revision_recommended"] = str(state["a4_revision_recommended"])
                    final_row["prompt_tokens"] = str(
                        state["a1_prompt_tokens"] + state["a2_prompt_tokens"]
                        + state["a3_prompt_tokens"] + state["a4_prompt_tokens"]
                    )
                    final_row["completion_tokens"] = str(
                        state["a1_completion_tokens"] + state["a2_completion_tokens"]
                        + state["a3_completion_tokens"] + state["a4_completion_tokens"]
                    )
                    final_row["wall_seconds"] = f"{state['wall_seconds']:.3f}"
                    final_row["worker_id"] = str(state["worker_id"])

                    results_buffer[row_index] = final_row
                    flush_buffer(writer, outfile)

                    total_a1_pt += state["a1_prompt_tokens"]
                    total_a1_ct += state["a1_completion_tokens"]
                    total_a1_errors += state["a1_errors"]
                    total_a2_pt += state["a2_prompt_tokens"]
                    total_a2_ct += state["a2_completion_tokens"]
                    total_a2_errors += state["a2_errors"]
                    total_a3_pt += state["a3_prompt_tokens"]
                    total_a3_ct += state["a3_completion_tokens"]
                    total_a3_errors += state["a3_errors"]
                    total_a4_pt += state["a4_prompt_tokens"]
                    total_a4_ct += state["a4_completion_tokens"]
                    total_a4_errors += state["a4_errors"]
                    total_a3_revisions_recommended += state["a3_revision_recommended"]
                    total_a4_revisions_recommended += state["a4_revision_recommended"]
                    total_sequential_time += state["wall_seconds"]

                    if result.get("a3_revision_recommended", 0):
                        identifier = (
                            state["row"].get("essay_id")
                            or state["row"].get("identifier")
                            or f"row_{row_index}"
                        ).strip()
                        a3_revision_ids.append(identifier)
                    if result.get("a4_revision_recommended", 0):
                        identifier = (state["row"].get("identifier") or f"row_{row_index}").strip()
                        a4_revision_ids.append(identifier)

                    if result["status"] == "skip":
                        identifier = (state["row"].get("identifier") or f"row_{row_index}").strip()
                        skipped_ids.append(identifier)

                    status_label = (
                        "ERROR (retry exhausted)"
                        if result["status"] == "error"
                        else "SKIPPED (content filter)"
                        if result["status"] == "skip"
                        else f"grade={result['grade']}"
                    )
                    print(
                        f"[{completed_rows:>4}/{len(rows_to_process)}] row={row_index:>4} "
                        f"worker={state['worker_id']} | {status_label} | "
                        f"target={target_workers} | wall={state['wall_seconds']:.1f}s | "
                        f"tokens A1 in={state['a1_prompt_tokens']} out={state['a1_completion_tokens']}  "
                        f"A2 in={state['a2_prompt_tokens']} out={state['a2_completion_tokens']}  "
                        f"A3 in={state['a3_prompt_tokens']} out={state['a3_completion_tokens']}  "
                        f"A4 in={state['a4_prompt_tokens']} out={state['a4_completion_tokens']} | "
                        f"errors A1={state['a1_errors']} A2={state['a2_errors']} "
                        f"A3={state['a3_errors']} A4={state['a4_errors']}"
                    )

        if results_buffer:
            raise RuntimeError("Output order buffer was not fully flushed.")

        run_end_time = datetime.now()
        run_duration = run_end_time - run_start_time
        duration_str = str(run_duration).split(".")[0]
        total_prompt = total_a1_pt + total_a2_pt + total_a3_pt + total_a4_pt
        total_completion = total_a1_ct + total_a2_ct + total_a3_ct + total_a4_ct
        total_tokens = total_prompt + total_completion
        total_wall_seconds = time.perf_counter() - wall_start
        speedup_factor = (
            total_sequential_time / total_wall_seconds if total_wall_seconds > 0 else 0.0
        )

        append_summary_to_excel(
            xlsx_path=SUMMARY_XLSX,
            run_start=run_start_time,
            run_end=run_end_time,
            duration_str=duration_str,
            sequential_time=total_sequential_time,
            total_wall_seconds=total_wall_seconds,
            speedup_factor=speedup_factor,
            test_name=THIS_DIR.name,
            total_a1_prompt_tokens=total_a1_pt,
            total_a1_completion_tokens=total_a1_ct,
            total_a2_prompt_tokens=total_a2_pt,
            total_a2_completion_tokens=total_a2_ct,
            total_a3_prompt_tokens=total_a3_pt,
            total_a3_completion_tokens=total_a3_ct,
            total_a4_prompt_tokens=total_a4_pt,
            total_a4_completion_tokens=total_a4_ct,
            total_a1_errors=total_a1_errors,
            total_a2_errors=total_a2_errors,
            total_a3_errors=total_a3_errors,
            total_a4_errors=total_a4_errors,
            total_tokens=total_tokens,
            skipped_count=len(skipped_ids),
            skipped_ids=skipped_ids,
            start_workers=start_workers,
            peak_workers=peak_target_workers,
            min_workers=min_target_workers,
            retryable_events=adaptive_retryable_events,
            error_rows=adaptive_error_rows,
            total_a3_revisions_recommended=total_a3_revisions_recommended,
            total_a4_revisions_recommended=total_a4_revisions_recommended,
            a3_revision_ids=a3_revision_ids,
            a4_revision_ids=a4_revision_ids,
        )

        # trailing summary row in CSV
        summary: dict[str, str] = {field: "" for field in output_fieldnames}
        summary["assignment"] = "=== RUN SUMMARY ==="
        summary["AI_reasoning"] = (
            f"start={run_start_time.strftime('%Y-%m-%d %H:%M:%S')} | "
            f"end={run_end_time.strftime('%Y-%m-%d %H:%M:%S')} | "
            f"duration={duration_str} | "
            f"sequential_time={total_sequential_time:.3f} | "
            f"total_wall_seconds={total_wall_seconds:.3f} | "
            f"speedup_factor={speedup_factor:.2f} | "
            f"workers={start_workers} | peak_workers={peak_target_workers} | "
            f"min_workers={min_target_workers} | "
            f"retryable_events={adaptive_retryable_events} | "
            f"error_rows={adaptive_error_rows} | "
            f"skipped={len(skipped_ids)} | "
            f"skipped_ids={', '.join(skipped_ids) if skipped_ids else 'none'}"
        )
        summary["a1_prompt_tokens"] = str(total_a1_pt)
        summary["a1_completion_tokens"] = str(total_a1_ct)
        summary["a2_prompt_tokens"] = str(total_a2_pt)
        summary["a2_completion_tokens"] = str(total_a2_ct)
        summary["a3_prompt_tokens"] = str(total_a3_pt)
        summary["a3_completion_tokens"] = str(total_a3_ct)
        summary["a4_prompt_tokens"] = str(total_a4_pt)
        summary["a4_completion_tokens"] = str(total_a4_ct)
        summary["prompt_tokens"] = str(total_prompt)
        summary["completion_tokens"] = str(total_completion)
        summary["wall_seconds"] = f"{total_wall_seconds:.3f}"
        summary["sequential_time"] = f"{total_sequential_time:.3f}"
        summary["total_wall_seconds"] = f"{total_wall_seconds:.3f}"
        summary["speedup_factor"] = f"{speedup_factor:.2f}"
        summary["adaptive_start_workers"] = str(start_workers)
        summary["adaptive_peak_workers"] = str(peak_target_workers)
        summary["adaptive_min_workers"] = str(min_target_workers)
        summary["adaptive_retryable_events"] = str(adaptive_retryable_events)
        summary["adaptive_error_rows"] = str(adaptive_error_rows)
        summary["a3_revision_recommended"] = str(total_a3_revisions_recommended)
        summary["a4_revision_recommended"] = str(total_a4_revisions_recommended)
        summary["a3_revision_recommended_ids"] = ", ".join(a3_revision_ids)
        summary["a4_revision_recommended_ids"] = ", ".join(a4_revision_ids)
        writer.writerow(summary)

    print("\n" + "=" * 60)
    print("AUTOSCORE RUN SUMMARY")
    print("=" * 60)
    print(f"  Start time              : {run_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  End time                : {run_end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Duration                : {duration_str}")
    print(f"  Sequential time         : {total_sequential_time:.3f}s")
    print(f"  Total wall seconds      : {total_wall_seconds:.3f}s")
    print(f"  Speedup factor          : {speedup_factor:.2f}x")
    print(f"  Start workers           : {start_workers}")
    print(f"  Peak workers            : {peak_target_workers}")
    print(f"  Min workers             : {min_target_workers}")
    print(f"  Retryable Azure events  : {adaptive_retryable_events}")
    print(f"  Error rows              : {adaptive_error_rows}")
    print(f"  A3 revisions recommended: {total_a3_revisions_recommended}")
    print(f"  A4 revisions recommended: {total_a4_revisions_recommended}")
    print(f"  Agent 1 prompt tokens   : {total_a1_pt:,}")
    print(f"  Agent 1 completion tok. : {total_a1_ct:,}")
    print(f"  Agent 2 prompt tokens   : {total_a2_pt:,}")
    print(f"  Agent 2 completion tok. : {total_a2_ct:,}")
    print(f"  Agent 3 prompt tokens   : {total_a3_pt:,}")
    print(f"  Agent 3 completion tok. : {total_a3_ct:,}")
    print(f"  Agent 4 prompt tokens   : {total_a4_pt:,}")
    print(f"  Agent 4 completion tok. : {total_a4_ct:,}")
    print(f"  -- Total prompt tokens  : {total_prompt:,}")
    print(f"  -- Total compl. tokens  : {total_completion:,}")
    print(f"  -- GRAND TOTAL tokens   : {total_tokens:,}")
    print("=" * 60)
    print(f"Output written to {OUTPUT_CSV}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())