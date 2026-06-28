import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from pathlib import Path
from itertools import combinations
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile

# ---------------------------------------------------------------------------
# Paths — add as many as you like; all pairs are tested automatically
# ---------------------------------------------------------------------------
PATHS = [
   
    '/home/simao/ThesisTesing/ASAP2.0/MAS-H1/H1bcontra/MASGraded.csv',
     '/home/simao/ThesisTesing/ASAP2.0/MAS-H1/H1bcontra2/MASGraded.csv',
]

# Column names
HUMAN_COL     = "score"
AI_COL        = "AI_grade"
ID_COL        = "essay_id"
WORDCOUNT_COL = "essay_word_count"
SENTINEL      = -1

# Word-count buckets
BUCKETS = [
    ("< 25quatile",                    None, 247),
    ("> 25quartile and < 75 quartile", 247,  445),
    ("> 75quartile",                   445,  None),
]

ALPHA = 0.05
EXCEL_OUTPUT = Path('/home/simao/ThesisTesing/QWK_ResultsASAP.xlsx')

# ---------------------------------------------------------------------------
# Load CSVs
# ---------------------------------------------------------------------------
dfs   = {Path(p).parent.name: pd.read_csv(p, sep=',') for p in PATHS}
names = list(dfs.keys())
pairs = list(combinations(names, 2))

print(f"Loaded {len(names)} conditions: {', '.join(names)}")
print(f"Running {len(pairs)} pairwise comparison(s), "
      f"each split into {len(BUCKETS)} word-count buckets\n")

# ---------------------------------------------------------------------------
# QWK helpers
# ---------------------------------------------------------------------------
def qwk(y_true, y_pred):
    return cohen_kappa_score(y_true, y_pred, weights='quadratic')

def bootstrap_qwk_diff_paired(true_v1, pred_v1, true_v2, pred_v2,
                               n_boot=10000, seed=123):
    rng  = np.random.default_rng(seed)
    n    = len(true_v1)
    assert len(true_v2) == n, "Paired test requires equal n"

    obs_v1   = qwk(true_v1, pred_v1)
    obs_v2   = qwk(true_v2, pred_v2)
    obs_diff = obs_v1 - obs_v2

    boot_diffs = []
    for _ in range(n_boot):
        idx = rng.integers(0, n, size=n)
        try:
            boot_diffs.append(
                qwk(true_v1[idx], pred_v1[idx]) -
                qwk(true_v2[idx], pred_v2[idx])
            )
        except Exception:
            continue

    boot_diffs = np.array(boot_diffs)
    shifted    = boot_diffs - np.mean(boot_diffs)
    p_value    = np.mean(np.abs(shifted) >= np.abs(obs_diff))
    ci_low, ci_high = np.percentile(boot_diffs, [2.5, 97.5])
    return obs_v1, obs_v2, obs_diff, p_value, ci_low, ci_high

# ---------------------------------------------------------------------------
# Holm-Bonferroni correction
# ---------------------------------------------------------------------------
def holm_bonferroni(p_values, alpha=0.05):
    m       = len(p_values)
    indexed = sorted(enumerate(p_values), key=lambda x: x[1])
    reject  = [False] * m
    adj_p   = [0.0]   * m
    stop    = False

    for rank, (orig_idx, p) in enumerate(indexed, start=1):
        adjusted = min(p * (m - rank + 1), 1.0)
        if rank > 1:
            adjusted = max(adjusted, adj_p[indexed[rank - 2][0]])
        adj_p[orig_idx] = adjusted

        if not stop and p <= alpha / (m - rank + 1):
            reject[orig_idx] = True
        else:
            stop = True

    return reject, adj_p

# ---------------------------------------------------------------------------
# Data preparation helpers
# ---------------------------------------------------------------------------
def prepare_df(df):
    df = df.copy()
    df[ID_COL]        = df[ID_COL].astype(str).str.strip()
    df[HUMAN_COL]     = pd.to_numeric(df[HUMAN_COL],     errors="coerce")
    df[AI_COL]        = pd.to_numeric(df[AI_COL],        errors="coerce")
    df[WORDCOUNT_COL] = pd.to_numeric(df[WORDCOUNT_COL], errors="coerce")

    missing_mask = df[HUMAN_COL].isna() | df[AI_COL].isna()
    n_missing    = missing_mask.sum()

    df_valid = df[~missing_mask].copy()
    df_valid[HUMAN_COL] = df_valid[HUMAN_COL].astype(int)
    df_valid[AI_COL]    = df_valid[AI_COL].astype(int)

    sentinel_mask = df_valid[AI_COL] == SENTINEL
    n_sentinel    = sentinel_mask.sum()

    clean_df = df_valid[~sentinel_mask].copy()
    return clean_df, n_missing, n_sentinel


def align_pair(df_a, df_b):
    merged = df_a.merge(df_b, on=ID_COL, suffixes=("_a", "_b"), how="inner")
    if f"{WORDCOUNT_COL}_a" in merged.columns:
        merged[WORDCOUNT_COL] = merged[f"{WORDCOUNT_COL}_a"]
    n_only_a = len(set(df_a[ID_COL]) - set(df_b[ID_COL]))
    n_only_b = len(set(df_b[ID_COL]) - set(df_a[ID_COL]))
    return merged, n_only_a, n_only_b


def apply_bucket(df, low, high):
    mask = pd.Series([True] * len(df), index=df.index)
    if low  is not None: mask &= df[WORDCOUNT_COL] >= low
    if high is not None: mask &= df[WORDCOUNT_COL] <  high
    return df[mask]

# ---------------------------------------------------------------------------
# Collect all tests (pair × bucket)
# ---------------------------------------------------------------------------
all_tests   = []
all_pvalues = []

for name_a, name_b in pairs:
    clean_a, n_missing_a, n_sentinel_a = prepare_df(dfs[name_a])
    clean_b, n_missing_b, n_sentinel_b = prepare_df(dfs[name_b])

    merged, n_only_a, n_only_b = align_pair(clean_a, clean_b)

    n_skipped_a = n_missing_a + n_sentinel_a + n_only_a
    n_skipped_b = n_missing_b + n_sentinel_b + n_only_b
    

    print(f"\n{'#'*65}")
    print(f"  {name_a}  vs  {name_b}")
    print(f"{'#'*65}")
    print(f"\n  ── Row diagnostics ──")
    print(f"  {'Dataset':<30} {'NaN dropped':>12} {'Sentinel (-1)':>14} "
          f"{'Unique (unmatched)':>20} {'Total skipped':>14}")
    print(f"  {'─'*62}")
    print(f"  {name_a:<30} {n_missing_a:>12} {n_sentinel_a:>14} "
          f"{n_only_a:>20} {n_skipped_a:>14}")
    print(f"  {name_b:<30} {n_missing_b:>12} {n_sentinel_b:>14} "
          f"{n_only_b:>20} {n_skipped_b:>14}")
    print(f"\n  Matched rows used for comparison: {len(merged)}")

    if merged.empty:
        print("  ⚠  No overlapping rows — skipping.")
        continue

    for bucket_label, low, high in BUCKETS:
        sub = apply_bucket(merged, low, high)

        print(f"\n  ── Bucket: {bucket_label} ──  (n={len(sub)})")

        if len(sub) < 2:
            print("     ⚠  Not enough rows for QWK — skipping bucket.")
            continue

        true_a = sub[f"{HUMAN_COL}_a"].values
        pred_a = sub[f"{AI_COL}_a"].values
        true_b = sub[f"{HUMAN_COL}_b"].values
        pred_b = sub[f"{AI_COL}_b"].values

        qwk_a, qwk_b, diff, p_val, ci_low, ci_high = bootstrap_qwk_diff_paired(
            true_a, pred_a, true_b, pred_b
        )

        all_tests.append({
            "bucket":      bucket_label,
            "qwk_a":       qwk_a,
            "qwk_b":       qwk_b,
            "diff":        diff,
            "p_raw":       p_val,
            "ci_low":      ci_low,
            "ci_high":     ci_high,
            "n":           len(sub),
            "n_skipped_a": n_skipped_a,
            "n_skipped_b": n_skipped_b,
        })
        all_pvalues.append(p_val)

# ---------------------------------------------------------------------------
# Holm-Bonferroni correction across ALL tests
# ---------------------------------------------------------------------------
m = len(all_tests)
if m == 0:
    print("\nNo valid tests to run.")
else:
    reject_flags, adj_pvalues = holm_bonferroni(all_pvalues, alpha=ALPHA)
    for i, t in enumerate(all_tests):
        t["reject"] = reject_flags[i]
        t["p_adj"]  = adj_pvalues[i]

    # ── Console summary ────────────────────────────────────────────────────
    print(f"\n\n{'='*100}")
    print(f"  SUMMARY — Holm-Bonferroni corrected ({m} test(s), α = {ALPHA})")
    print(f"{'='*100}")
    print(f"  {'Bucket':<36} {'n':>5} {'QWK A':>7} {'QWK B':>7} "
          f"{'Diff':>7} {'p_raw':>7} {'p_adj':>7} {'Sig?':>5}")
    print(f"  {'─'*85}")
    for t in all_tests:
        sig = "YES" if t["reject"] else "no"
        print(f"  {t['bucket']:<36} {t['n']:>5} "
              f"{t['qwk_a']:>7.4f} {t['qwk_b']:>7.4f} {t['diff']:>+7.4f} "
              f"{t['p_raw']:>7.4f} {t['p_adj']:>7.4f} {sig:>5}")

    # ── Excel export ───────────────────────────────────────────────────────
    sheet_name = f"{name_a} vs {name_b}".replace(":", "-").replace("/", "-")[:31]

    if EXCEL_OUTPUT.exists():
        try:
            with zipfile.ZipFile(EXCEL_OUTPUT, 'r'):
                pass
            wb = load_workbook(EXCEL_OUTPUT)
        except zipfile.BadZipFile:
            print(f"⚠  Existing file is corrupt — recreating: {EXCEL_OUTPUT}")
            EXCEL_OUTPUT.unlink()
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(title=sheet_name)

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill("solid", start_color="2F4F7F")
    body_font   = Font(name="Arial", size=9)
    title_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_fill  = PatternFill("solid", start_color="1F3864")
    sig_fill    = PatternFill("solid", start_color="E2EFDA")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, font=None, fill=None, align=None, num_fmt=None):
        if font:  cell.font      = font
        if fill:  cell.fill      = fill
        if align: cell.alignment = align
        cell.border = border
        if num_fmt: cell.number_format = num_fmt

    headers = [
        ("Bucket",               28),
        ("n",                     6),
        (f"QWK A\n({names[0]})", 10),
        (f"QWK B\n({names[1]})", 10),
        ("Diff\n(A−B)",           9),
        ("95% CI",               18),
        ("p raw",                 8),
        ("p adj\n(Holm)",         9),
        ("Sig?",                  7),
    ]
    n_cols   = len(headers)
    last_col = get_column_letter(n_cols)

    # Skipped counts from first test (same for all buckets in a pair)
    n_skip_a = all_tests[0]["n_skipped_a"]
    n_skip_b = all_tests[0]["n_skipped_b"]

    # ── Title row ──
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = (f"{sheet_name}  |  "
                    f"n: {len(merged)}  |  "
                    f"α={ALPHA}")
    title_cell.font      = title_font
    title_cell.fill      = title_fill
    title_cell.alignment = center
    title_cell.border    = border
    ws.row_dimensions[1].height = 20

    # ── Column headers ──
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        style_cell(cell, font=header_font, fill=header_fill, align=center)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28

    # ── Data rows ──
    row = 3
    for t in all_tests:
        is_sig = t["reject"]
        values = [
            t["bucket"],
            t["n"],
            t["qwk_a"],
            t["qwk_b"],
            t["diff"],
            f"[{t['ci_low']:.4f}, {t['ci_high']:.4f}]",
            t["p_raw"],
            t["p_adj"],
            "YES ✓" if is_sig else "no",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            fill = sig_fill if is_sig else None
            style_cell(cell, font=body_font, fill=fill, align=center)

            if col_idx in (3, 4):
                cell.number_format = "0.0000"
            elif col_idx == 5:
                cell.number_format = "+0.0000;-0.0000;0.0000"
            elif col_idx in (7, 8):
                cell.number_format = "0.0000"

        ws.row_dimensions[row].height = 14
        row += 1

    # ── Freeze panes & auto-filter ──
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}2"

    # ── Metadata footer ──
    row += 1
    ws.cell(row=row, column=1, value="Paths used:").font = Font(name="Arial", bold=True, size=9)
    for i, p in enumerate(PATHS, start=1):
        ws.cell(row=row + i, column=1, value=p).font = Font(name="Arial", size=9, italic=True)

    wb.save(EXCEL_OUTPUT)
    print(f"\n✓  Results appended to: {EXCEL_OUTPUT}  (sheet: '{sheet_name}')")

print(f"\nPaths used:\n" + "\n".join(f"  {p}" for p in PATHS))